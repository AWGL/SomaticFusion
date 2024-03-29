import pandas
from openpyxl import Workbook
import os
import numpy
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
from openpyxl.styles.borders import Border, Side, BORDER_MEDIUM, BORDER_THIN, BORDER_THICK
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment




def get_referral_list(referral, referrals_path, wb):
	
	#get the genes in the referral file for the sample
	referral_file=pandas.read_csv(referrals_path+"Referrals/"+referral+".txt", sep="\t")
	referral_list=list(referral_file['Genes'])
	return(referral_list, wb)


def get_NTC_fusion_report(ntc, referral_list, path,wb, ws6):
	
	#get the NTC star fusion results for all the genes in the samples referral type
	ws6["A4"]="NTC STAR-Fusion results"
	for referral in referral_list:
                
		if ((referral!= "MET_exon14_skipping") and (referral!="EGFRv3")):
			if (os.stat(path + ntc+"/Results/STAR_Fusion/"+ntc+"_fusion_report_"+referral+".txt").st_size!=0):
      	
				NTC_star_fusion_report=pandas.read_csv(path +ntc+"/Results/STAR_Fusion/"+ntc+"_fusion_report_"+referral+".txt", sep="\t")
	
				NTC_star_fusion_report=pandas.DataFrame(NTC_star_fusion_report)

				#if there are no fusions for the gene in the report append "<gene name>- no fusions found"
				if (len(NTC_star_fusion_report)==0):
					dict_NTC_star_fusion=({"Fusion_name": referral + "-no fusions found", "Split_Read_Count": "", "Spanning_Read_Count": "", "Left_Breakpoint": "", "Right_Breakpoint": "", "SpliceType": "", "LargeAnchorSupport": "", "FFPM": "", "LeftBreakEntropy": "", "RightBreakEntropy": "", "CDS_Left_ID": "", "CDS_Left_Range": "", "CDS_Right_ID": "", "CDS_Right_Range": "", "Prot_Fusion_Type": "", "Num_WT_Fragments_Left": "", "Num_WT_Fragments_Right": "", "Fusion_Allelic_Fraction": ""}) 
					NTC_star_fusion_report= pandas.DataFrame(dict_NTC_star_fusion, columns=["Fusion_name", "Split_Read_Count", "Spanning_Read_Count", "Left_Breakpoint", "Right_Breakpoint", "SpliceType", "LargeAnchorSupport", "FFPM", "type", "LeftBreakEntropy", "RightBreakEntropy", "CDS_Left_ID", "CDS_Left_Range", "CDS_Right_ID", "CDS_Right_Range", "Prot_Fusion_Type", "Num_WT_Fragments_Left", "Num_WT_Fragments_Right", "Fusion_Allelic_Fraction"], index=[0]) 	

				else:
					del NTC_star_fusion_report['Unnamed: 0']

				for row in dataframe_to_rows(NTC_star_fusion_report, header=False, index=False):
					ws6.append(row)

			else:
				print("NTC STAR-Fusion results cannot be found")


	#get the NTC arriba results for all the genes in the samples referral type
	ws6["A19"]="Arriba results"
	for referral in referral_list:

		if ((referral!= "MET_exon14_skipping") and (referral!="EGFRv3")):
	
			if (os.stat(path+ntc+"/Results/arriba/"+ntc+"_fusion_report_"+referral+"_arriba.txt").st_size!=0):    	
				NTC_arriba_report=pandas.read_csv(path +ntc+"/Results/arriba/"+ntc+"_fusion_report_"+referral+"_arriba.txt", sep="\t")
				#if there are no fusions for the gene in the report append "<gene name>- no fusions found"
				if (len(NTC_arriba_report)==0):
					dict_NTC_arriba=({"gene1": referral + "-no fusions found", "gene2": "", "strand1(gene/fusion)": "", "strand2(gene/fusion)": "", "breakpoint1": "", "breakpoint2": "", "site1": "", "site2": "", "type": "", "direction1": "", "direction2": "", "split_reads1": "", "split_reads2": "", "discordant_mates": "", "coverage1": "", "coverage2": "", "confidence": "", "closest_genomic_breakpoint1": "", "closest_genomic_breakpoint2": "", "filters": "", "fusion_transcript": "", "reading_frame" : "", "peptide_sequence": "", "read_identifiers": ""}) 
					NTC_arriba_report= pandas.DataFrame(dict_NTC_arriba, columns=["gene1", "gene2", "strand1(gene/fusion)", "strand2(gene/fusion)", "breakpoint1", "breakpoint2", "site1", "site2", "type", "direction1", "direction2", "split_reads1", "split_reads2", "discordant_mates", "coverage1", "coverage2", "confidence", "closest_genomic_breakpoint1", "closest_genomic_breakpoint2", "filters", "fusion_transcript", "reading_frame", "peptide_sequence", "read_identifiers"], index=[0]) 	
				else:
					del NTC_arriba_report['Unnamed: 0']	
				for row in dataframe_to_rows(NTC_arriba_report, header=False, index=False):
						ws6.append(row)
			else:
				print("NTC arriba file does not exist")


	return(NTC_star_fusion_report, NTC_arriba_report, wb, ws6)

	

def get_star_fusion_report(referral_list, ntc_star_fusion_report, sampleId, path, wb, ws1, ws5):
	ws1["A9"]="STAR-Fusion results"

	#create empty pandas dataframe and append the STAR-Fusion results for all the genes in the sample's referral type

	star_fusion_report_final= pandas.DataFrame(columns=["Fusion_Name", "Split_Read_Count", "Spanning_Read_Count", "Left_Breakpoint", "Right_Breakpoint", "SpliceType", "LargeAnchorSupport", "FFPM", "LeftBreakEntropy", "RightBreakEntropy","CDS_Left_ID", "CDS_Left_Range", "CDS_Right_ID", "CDS_Right_Range", "Prot_Fusion_Type","Num_WT_Fragments_Left", "Num_WT_Fragments_Right", "Fusion_Allelic_Fraction"]) 

	for referral in referral_list:

		if ((referral!= "MET_exon14_skipping") and (referral!="EGFRv3")):


			if (os.stat(path+sampleId+"/Results/STAR_Fusion/"+sampleId+"_fusion_report_"+referral+".txt").st_size!=0):
      
				star_fusion_report=pandas.read_csv(path+sampleId+"/Results/STAR_Fusion/"+sampleId+"_fusion_report_"+referral+".txt", sep="\t")	
				if len(star_fusion_report!=0):
					star_fusion_report=pandas.DataFrame(star_fusion_report)
					del star_fusion_report['Unnamed: 0']
				else:
					#if there are no fusions for the gene in the report append "<gene name>- no fusions found"
					dict_star_fusion=({"Fusion_Name": referral+"-no fusions found", "Split_Read_Count": "", "Spanning_Read_Count": "", "Left_Breakpoint": "", "Right_Breakpoint": "", "SpliceType": "", "LargeAnchorSupport": "", "FFPM": "", "LeftBreakEntropy": "", "RightBreakEntropy": "","CDS_Left_ID": "", "CDS_Left_Range": "", "CDS_Right_ID": "", "CDS_Right_Range": "", "Prot_Fusion_Type": "","Num_WT_Fragments_Left": "", "Num_WT_Fragments_Right": "", "Fusion_Allelic_Fraction": ""}) 
					star_fusion_report= pandas.DataFrame(dict_star_fusion, columns=["Fusion_Name", "Split_Read_Count", "Spanning_Read_Count", "Left_Breakpoint", "Right_Breakpoint", "SpliceType", "LargeAnchorSupport", "FFPM", "LeftBreakEntropy", "RightBreakEntropy","CDS_Left_ID", "CDS_Left_Range", "CDS_Right_ID", "CDS_Right_Range", "Prot_Fusion_Type","Num_WT_Fragments_Left", "Num_WT_Fragments_Right", "Fusion_Allelic_Fraction"], index=[0]) 	

				star_fusion_report_final=star_fusion_report_final.append(star_fusion_report, sort=False)

			else:
				print("STAR-Fusion report does not exist")
	

	for row in dataframe_to_rows(star_fusion_report_final, header=True, index=False):
		ws1.append(row)	
	#Show an error message if not all the fusions will fit on the analysis sheet
	if (len(star_fusion_report_final)>12) :
		ws1["A22"]="ERROR:NOT ENOUGH SPACE FOR ALL FUSIONS-CHECK RESULTS FILES"
	if (len(star_fusion_report_final)>12) :
		ws1["A22"]="ERROR:NOT ENOUGH SPACE FOR ALL FUSIONS-CHECK RESULTS FILES"
		colour= PatternFill("solid", fgColor="00FF0000")
		position= ['A22']
		for cell in position:
			ws1[cell].fill=colour
		position= ['A20']
		for cell in position:
			ws5[cell].fill=colour
		

	return(star_fusion_report_final, wb, ws1, ws5)		
		


def get_arriba_fusion_report(referral_list, sampleId, path, wb, ws1,ws5):

	ws1["A23"]="Arriba results"


	#create empty pandas dataframe and append the arriba results for all the genes in the sample's referral type

	arriba_report_final= pandas.DataFrame(columns=["gene1", "gene2", "strand1(gene/fusion)", "strand2(gene/fusion)", "breakpoint1", "breakpoint2", "site1", "site2", "type", "direction1", "direction2", "split_reads1", "split_reads2", "discordant_mates", "coverage1", "coverage2", "confidence", "closest_genomic_breakpoint1", "closest_genomic_breakpoint2", "filters", "fusion_transcript", "reading_frame", "peptide_sequence", "read_identifiers"])

	for referral in referral_list:

		if ((referral!= "MET_exon14_skipping") and (referral!="EGFRv3")):

			if (os.stat(path+ sampleId+"/Results/arriba/"+sampleId+"_fusion_report_"+referral+"_arriba.txt").st_size!=0):     	
				arriba_report=pandas.read_csv(path+ sampleId+"/Results/arriba/"+sampleId+"_fusion_report_"+referral+"_arriba.txt", sep="\t")

				if len(arriba_report!=0):
					del arriba_report['Unnamed: 0']					

				else:
					#if there are no fusions for the gene in the report append "<gene name>- no fusions found"
					dict_arriba=({"gene1": referral + "-no fusions found", "gene2": "", "strand1(gene/fusion)": "", "strand2(gene/fusion)": "", "breakpoint1": "", "breakpoint2": "", "site1": "", "site2": "", "type": "", "direction1": "", "direction2": "", "split_reads1": "", "split_reads2": "", "discordant_mates": "", "coverage1": "", "coverage2": "", "confidence": "", "closest_genomic_breakpoint1": "", "closest_genomic_breakpoint2": "", "filters": "", "fusion_transcript": "", "reading_frame" : "", "peptide_sequence": "", "read_identifiers": ""}) 
					arriba_report= pandas.DataFrame(dict_arriba, columns=["gene1", "gene2", "strand1(gene/fusion)", "strand2(gene/fusion)", "breakpoint1", "breakpoint2", "site1", "site2", "type", "direction1", "direction2", "split_reads1", "split_reads2", "discordant_mates", "coverage1", "coverage2", "confidence", "closest_genomic_breakpoint1", "closest_genomic_breakpoint2", "filters", "fusion_transcript", "reading_frame", "peptide_sequence", "read_identifiers"], index=[0]) 	

				arriba_report_final=arriba_report_final.append(arriba_report, sort=False)

			else:
				print("arriba report does not exist")

	for row in dataframe_to_rows(arriba_report_final, header=True, index=False):
		ws1.append(row)
	#Show an error message if not all the fusions will fit on the analysis sheet
	if (len(arriba_report_final)>12) :
		ws1["A36"]="ERROR:NOT ENOUGH SPACE FOR ALL FUSIONS-CHECK RESULTS FILES"
		colour= PatternFill("solid", fgColor="00FF0000")
		position= ['A36']
		for cell in position:
			ws1[cell].fill=colour
		position= ['A34']
		for cell in position:
			ws5[cell].fill=colour
				

	return(arriba_report_final, wb, ws1, ws5)


def get_ntc_total_coverage(ntc ,path):

	#get the ntc total coverage with and without duplicates 

	if (os.stat(path+ntc+"/"+ntc+"_coverage.totalCoverage").st_size!=0):
		ntc_total_coverage=pandas.read_csv(path+ntc+"/"+ntc+"_coverage.totalCoverage", sep="\t")
		ntc_total_average_depth=ntc_total_coverage["AVG_DEPTH"]
	else:
		print("NTC total coverage file does not exist")

	if (os.stat(path+ntc+"/"+ntc+"_rmdup_coverage.totalCoverage").st_size!=0):
		ntc_total_coverage_rmdup=pandas.read_csv(path+ntc+"/"+ntc+"_rmdup_coverage.totalCoverage", sep="\t")
		ntc_total_average_depth_rmdup=ntc_total_coverage_rmdup["AVG_DEPTH"]
	else:
		print("NTC total coverage rmdup file does not exist")

	return(ntc_total_average_depth, ntc_total_average_depth_rmdup)


def get_total_coverage(ntc_total_average_depth, ntc_total_average_depth_rmdup, sampleId, path, wb,ws2):

	#get the total coverage of the genes of interest with and without duplicates 
	ws2["A2"]="with duplicates"
	if (os.stat(path+sampleId+"/"+sampleId+"_coverage.totalCoverage").st_size!=0):
		total_coverage=pandas.read_csv(path+sampleId+"/"+sampleId+"_coverage.totalCoverage", sep="\t")
		total_coverage["NTC_AVG_DEPTH"]=ntc_total_average_depth
		total_coverage["%NTC contamination"]=((total_coverage["NTC_AVG_DEPTH"].div(total_coverage["AVG_DEPTH"]).fillna(0)))*100
		del total_coverage["PERC_COVERAGE@100"]
		for row in dataframe_to_rows(total_coverage, header=True, index=False):
			ws2.append(row)
	else:
		print(" Total coverage file does not exist")

	if (os.stat(path+sampleId+"/"+sampleId+"_rmdup_coverage.totalCoverage").st_size!=0):
		ws2["A32"]="without duplicates"
		total_coverage_rmdup=pandas.read_csv(path+sampleId+"/"+sampleId+"_rmdup_coverage.totalCoverage", sep="\t")
		total_coverage_rmdup["NTC_AVG_DEPTH"]=ntc_total_average_depth_rmdup
		total_coverage_rmdup["%NTC contamination"]=((total_coverage_rmdup["NTC_AVG_DEPTH"].div(total_coverage_rmdup["AVG_DEPTH"])).fillna(0))*100
		del total_coverage_rmdup["PERC_COVERAGE@100"]
		for row in dataframe_to_rows(total_coverage_rmdup, header=True, index=False):
			ws2.append(row)


		#highlight the cells in the percentage NTC contamination column red if the value is greater than 10%
		colour= PatternFill("solid", start_color="00FF0000", end_color="00FF0000")
		ws2.conditional_formatting.add('D4:D28', CellIsRule( operator='greaterThan', formula=['10'], stopIfTrue=True, fill=colour))


		#highlight the cells in the percentage NTC contamination column red if the value is greater than 10%
		colour= PatternFill("solid", start_color="00FF0000", end_color="00FF0000")
		ws2.conditional_formatting.add('D34:D56', CellIsRule( operator='greaterThan', formula=['10'], stopIfTrue=True, fill=colour))


	else:
		print(" Total coverage file rmdup does not exist")
	ws2.column_dimensions['A'].width=50

	return (total_coverage, total_coverage_rmdup, ws2, wb)


def ntc_get_coverage(path, ntc):

	#get the coverage of the regions of interest with and without duplicates in the ntc
	if (os.stat(path+ntc+"/"+ntc+"_coverage.coverage").st_size!=0):
		ntc_coverage=pandas.read_csv(path+ntc+"/"+ntc+"_coverage.coverage", sep="\t")
		ntc_average_depth=ntc_coverage["AVG_DEPTH"]
		del ntc_coverage["PERC_COVERAGE@100"]
	else:
		print(" NTC coverage file does not exist")

	if (os.stat(path+ntc+"/"+ntc+"_rmdup_coverage.coverage").st_size!=0):
		ntc_coverage_rmdup=pandas.read_csv(path+ntc+"/"+ntc+"_rmdup_coverage.coverage", sep="\t")
		ntc_average_depth_rmdup=ntc_coverage_rmdup["AVG_DEPTH"]
		del ntc_coverage_rmdup["PERC_COVERAGE@100"]
	else:
		print("NTC coverage rmdup file does not exist")

	return(ntc_average_depth, ntc_average_depth_rmdup)


def get_coverage(referral, ntc_average_depth, ntc_average_depth_rmdup, sampleId, path, wb, ws3, ws4):

	#get the coverage of the regions of interest with and without duplicates in the sample
	if (os.stat(path+sampleId+"/"+sampleId+"_coverage.coverage").st_size!=0):
		coverage=pandas.read_csv(path+sampleId+"/"+sampleId+"_coverage.coverage", sep="\t")
		coverage["NTC_AVG_DEPTH"]=ntc_average_depth
		coverage["%NTC contamination"]=((coverage["NTC_AVG_DEPTH"].div(coverage["AVG_DEPTH"])).fillna(0))*100
		del coverage["PERC_COVERAGE@100"]
		for row in dataframe_to_rows(coverage, header=True, index=False):
			ws3.append(row)

		#highlight the cells in the percentage NTC contamination column red if the value is greater than 10%
		colour= PatternFill("solid", start_color="00FF0000", end_color="00FF0000")
		ws3.conditional_formatting.add('G2:G300', CellIsRule( operator='greaterThan', formula=['10'], stopIfTrue=True, fill=colour))
	else:
		print("Coverage file does not exist")


	if (os.stat(path+sampleId+"/"+sampleId+"_rmdup_coverage.coverage").st_size!=0):
		coverage_rmdup=pandas.read_csv(path+sampleId+"/"+sampleId+"_rmdup_coverage.coverage", sep="\t")
		coverage_rmdup["NTC_AVG_DEPTH"]=ntc_average_depth_rmdup
		coverage_rmdup["%NTC contamination"]=((coverage_rmdup["NTC_AVG_DEPTH"].div(coverage_rmdup["AVG_DEPTH"])).fillna(0))*100
		del coverage_rmdup["PERC_COVERAGE@100"]
		for row in dataframe_to_rows(coverage_rmdup, header=True, index=False):
			ws4.append(row)

		#highlight the cells in the percentage NTC contamination column red if the value is greater than 10%
		colour= PatternFill("solid", start_color="00FF0000", end_color="00FF0000")
		ws4.conditional_formatting.add('G2:G300', CellIsRule( operator='greaterThan', formula=['10'], stopIfTrue=True, fill=colour))
	else:
		print("Coverage rmdup file does not exist")


	ws3.column_dimensions['B'].width=15
	ws4.column_dimensions['B'].width=15
	ws3.column_dimensions['C'].width=15
	ws4.column_dimensions['C'].width=15
	ws3.column_dimensions['D'].width=110
	ws4.column_dimensions['D'].width=110


	return(coverage_rmdup, coverage, ws3, ws4, wb)



def format_analysis_sheet(referral, wb,ws1, ws2, ws3,ws4,ws5,ws6,ws7,ws9):


	ws1['A5']=sampleId
	ws1['B5']='=Patient_demographics!C2'
	ws1['C5']=referral
	ws1['D5']=worksheetId
	ws1['E5']=seqId
	ws1['C1']="Analysis sheet- Pan RNA Cancer Fusion Panel"


	ws1['S10']="Conclusion checker 1"
	ws1['T10']="Conclusion checker 2"
	ws1['U10']="Comments"
	ws1['Y24']="Conclusion checker 1"
	ws1['Z24']="Conclusion checker 2"
	ws1['AA24']="Comments"

	dv = DataValidation(type="list", formula1='"Genuine,Artefact,Fail_QC,no fusions"', allow_blank=True)
	ws1.add_data_validation(dv)

	dv.add('S11:S21')
	dv.add('T11:T21')
	dv.add('Y25:Y35')
	dv.add('Z21:Z35')

	font_bold=Font(bold=True)
	position= ['A4', 'B4', 'C4', 'D4', 'E4', 'F4', 'G4', 'A39',
	'A10','B10','C10','D10','E10','F10','G10','H10','I10','J10','K10','L10','M10','N10','O10','P10','Q10','R10','S10','T10', 'U10',
	'A24','B24','C24','D24','E24','F24','G24','H24','I24','J24','K24','L24','M24','N24','O24','P24','Q24','R24','S24','T24','U24','V24','W24','X24','Y24', 'Z24', 'AA24']
	for cell in position:
		ws1[cell].font=font_bold


	for column in ['A','B', 'Q']:
		ws5.column_dimensions[column].width=25

	for column in ['A','B', 'C','D','E', 'F','G', 'H', 'I', 'J','K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y']:
		ws1.column_dimensions[column].width=20	

	border_a=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
	position=['A4','B4','C4','D4','E4','F4','G4','A5','B5','C5','D5','E5','F5','G5','A10','B10','C10','D10','E10','F10','G10','H10','I10','J10','K10','L10','M10','N10','O10','P10','Q10','R10','S10','T10','U10',
	'A24','B24','C24','D24','E24','F24','G24','H24','I24','J24','K24','L24','M24','N24','O24','P24','Q24','R24','S24','T24','U24','V24','W24','X24','Y24', 'Z24', 'AA24']
	for cell in position:
		ws1[cell].border=border_a

	colour= PatternFill("solid", fgColor="DCDCDC")
	position= ['A10','B10','C10','D10','E10','F10','G10','H10','I10','J10','K10','L10','M10','N10','O10','P10','Q10','R10',
	'A24','B24','C24','D24','E24','F24','G24','H24','I24','J24','K24','L24','M24','N24','O24','P24','Q24','R24','S24','T24','U24','V24','W24','X24']
	for cell in position:
		ws1[cell].fill=colour


	colour= PatternFill("solid", fgColor="00FFCC00")
	position= ['F4', 'G4', 'S10', 'T10', 'U10', 'Y24', 'Z24', 'AA24']
	for cell in position:
		ws1[cell].fill=colour

	colour= PatternFill("solid", fgColor="00CCFFFF")
	position= ['A4','B4','C4','D4','E4']
	for cell in position:
		ws1[cell].fill=colour



	ws5['C1']='Analysis Sheet-Pan RNA Cancer Fusion Panel'
	ws5['C1'].font=Font(size=16)
	ws5["A3"]="Lab number" 
	ws5["B3"]="Patient name"
	ws5["C3"]="Tumour %"
	ws5["D3"]="Reason for Referral"
	ws5["E3"]="Qubit RNA conc after pre-processing (ng/ul)"
	ws5["F3"]="DV200"
	ws5["G3"]="% mapped reads-duplicates removed"
	ws5["H3"]="Due date"

	ws5["I3"]="NGS wks"
	ws5["J3"]="NextSeq run ID"
	ws5["K3"]="NTC check 1 result"
	ws5["L3"]="NTC check 2 result"
	ws5["M3"]="Analysed by:"
	ws5["N3"]="Checked by:"


	ws5["A6"]= "Gene fusion genuine calls"
	ws5["A7"]= "STAR-Fusion results"
	
	ws5["A8"]="Fusion_Name"
	ws5["B8"]="Split_Read_Count"
	ws5["C8"]="Spanning_Read_Count"
	ws5["D8"]="Left_Breakpoint"
	ws5["E8"]="Right_Breakpoint:"
	ws5["F8"]="SpliceType"
	ws5["G8"]="LargeAnchorSupport"
	ws5["H8"]="FFPM"
	ws5["I8"]="Prot_Fusion_Type"
	ws5["J8"]="Fusion_Allelic_Fraction"
	ws5["O8"]="Conclusion checker 1"
	ws5["P8"]="Conclusion checker 2"
	ws5["Q8"]="Comments"
	
	ws5["A21"]="Arriba results"
	ws5["A22"]="#gene1"
	ws5["B22"]="gene2"
	ws5["C22"]="strand1(gene/fusion)"
	ws5["D22"]="strand2(gene/fusion)"
	ws5["E22"]="breakpoint1"
	ws5["F22"]="breakpoint2"
	ws5["G22"]="site1"
	ws5["H22"]="site2"
	ws5["I22"]="type"
	ws5["J22"]="split_reads1"
	ws5["K22"]="split_reads2"
	ws5["L22"]="discordant_mates"
	ws5["M22"]="confidence"
	ws5["N22"]="filters"
	ws5["O22"]="Conclusion checker 1"
	ws5["P22"]="Conclusion checker 2"
	ws5["Q22"]="Comments"

	ws5["A36"]="Gene fusions quality metrics"
	ws5["A38"]="Post PCR1 Qubit"
	ws5["A39"]="Total mapped reads-duplicates removed"
	ws5["B37"]="Value"
	ws5["C37"]="Conclusion checker 1"
	ws5["D37"]="Conclusion checker 2"


	#Increase the height of all rows in the summary report to allow each cell to contain two lines of text
	for row in range(55):
		ws5.row_dimensions[row].height=30



	
	border_a=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
	position=['A3','B3','C3','D3','E3','F3','G3', 'H3',
	'A4','B4','C4','D4','E4','F4', 'G4','H4',
        'I3','J3','K3','L3','M3','N3',
        'I4','J4','K4','L4','M4','N4',
        'A8','B8','C8','D8','E8','F8','G8', 'H8', 'I8', 'J8', 'K8', 'L8', 'M8', 'N8', 'O8', 'P8', 'Q8',
        'A22','B22','C22','D22','E22','F22','G22', 'H22', 'I22', 'J22', 'K22', 'L22', 'M22', 'N22', 'O22', 'P22', 'Q22',
        'B37', 'C37', 'D37',
        'A38', 'B38', 'C38', 'D38',
        'A39', 'B39', 'C39', 'D39']
	for cell in position:
		ws5[cell].border=border_a


	colour= PatternFill("solid", fgColor="DCDCDC")
	position= ['A8', 'B8', 'C8', 'D8', 'E8', 'F8', 'G8', 'H8', 'I8', 'J8', 'K8', 'L8', 'M8','N8',
                   'A22', 'B22', 'C22', 'D22', 'E22', 'F22', 'G22', 'H22', 'I22', 'J22', 'K22', 'L22' , 'M22', 'N22',
                   'B37', 'A38', 'A39']
	for cell in position:
		ws5[cell].fill=colour


	colour= PatternFill("solid", fgColor="00CCFFFF")
	position= ['A3','B3','C3','D3','E3','F3', 'G3', 'H3', 'I3','J3']
	for cell in position:
		ws5[cell].fill=colour



	colour= PatternFill("solid", fgColor="00FFCC00")
	position= ['K3', 'L3', 'M3', 'N3', 'O8', 'P8', 'Q8', 'O22', 'P22', 'Q22','C37', 'D37']
	for cell in position:
		ws5[cell].fill=colour


	font_bold=Font(bold=True)
	position= ['A3','B3','C3','D3','E3','F3','G3','H3','I3','J3','K3','L3','M3','N3','A8','B8','C8','D8','E8','F8','G8','H8','I8', 'J8', 'O8','P8', 'Q8',
		'A22','B22','C22','D22','E22','F22','G22','H22','I22', 'J22', 'K22', 'L22', 'M22', 'N22', 'O22', 'P22', 'Q22',
		'B37', 'C37', 'D37', 'A38', 'A39',
		'A42','B42','C42','D42','E42','F42','G42','H42','I42', 'J42', 'K42',
		'A47','B47','C47','D47','E47','F47','G47',
		'A6', 'A7', 'A21', 'A36', 'A41', 'A46']
	for cell in position:
		ws5[cell].font=font_bold


	ws5['A4']=sampleId
	ws5['B4']= '=Patient_demographics!C2'
	ws5['C4']= '=Patient_demographics!I2'
	ws5['D4']=referral
	ws5['E4']= '=Patient_demographics!L2'
	ws5['F4']= '=Patient_demographics!J2'
	ws5['H4']= '=Patient_demographics!H2'
	ws5['H4'].number_format='mm-dd-yy'


	ws5['I4']=worksheetId
	ws5['J4']=seqId
	ws5['K4']='=Subpanel_NTC_Check!I7'
	ws5['L4']='=Subpanel_NTC_Check!I8'
	ws5['B38']='=Patient_demographics!O2'


	#add pass, fail dropdown boxes to the summary sheet
	dv = DataValidation(type="list", formula1='"PASS,FAIL"', allow_blank=True)
	ws5.add_data_validation(dv)
	dv.add('C38:C39')
	dv.add('D38:D39')
	dv.add('F48:F51')
	dv.add('G48:G51')


	#transfer values in star-fusion table of gene fusion report to the summary tab

	summary_column_list=['A', 'B','C','D','E','F','G','H','I','J', 'O', 'P', 'Q']
	fusion_report_column_list=[ 'A','B','C','D','E','F','G','H','O', 'R', 'S', 'T', 'U']

	value=0
	while (value<(len(fusion_report_column_list))):
		fusion_report_column=fusion_report_column_list[value]
		summary_column=summary_column_list[value]

		ws5[summary_column+'9']= '=IF(OR(Gene_fusion_report!S11<>"Artefact",Gene_fusion_report!T11<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'11, "0")'
		ws5[summary_column+'10']= '=IF(OR(Gene_fusion_report!S12<>"Artefact",Gene_fusion_report!T12<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'12, "0")'
		ws5[summary_column+'11']= '=IF(OR(Gene_fusion_report!S13<>"Artefact",Gene_fusion_report!T13<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'13, "0")'
		ws5[summary_column+'12']= '=IF(OR(Gene_fusion_report!S14<>"Artefact",Gene_fusion_report!T14<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'14, "0")'
		ws5[summary_column+'13']= '=IF(OR(Gene_fusion_report!S15<>"Artefact",Gene_fusion_report!T15<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'15, "0")'
		ws5[summary_column+'14']= '=IF(OR(Gene_fusion_report!S16<>"Artefact",Gene_fusion_report!T16<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'16, "0")'
		ws5[summary_column+'15']= '=IF(OR(Gene_fusion_report!S17<>"Artefact",Gene_fusion_report!T17<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'17, "0")'
		ws5[summary_column+'16']= '=IF(OR(Gene_fusion_report!S18<>"Artefact",Gene_fusion_report!T18<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'18, "0")'
		ws5[summary_column+'17']= '=IF(OR(Gene_fusion_report!S19<>"Artefact",Gene_fusion_report!T19<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'19, "0")'
		ws5[summary_column+'18']= '=IF(OR(Gene_fusion_report!S20<>"Artefact",Gene_fusion_report!T20<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'20, "0")'
		ws5[summary_column+'19']= '=IF(OR(Gene_fusion_report!S21<>"Artefact",Gene_fusion_report!T21<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'21, "0")'
		ws5[summary_column+'20']= '=IF(OR(Gene_fusion_report!S22<>"Artefact",Gene_fusion_report!T22<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'22, "0")'

		value=value+1


	#transfer values in arriba table of gene fusion report to the summary tab

	summary_column_list=['A', 'B','C','D','E','F','G','H','I','J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']
	fusion_report_column_list=[ 'A','B','C','D','E','F','G','H','I', 'L', 'M', 'N', 'Q', 'T','Y','Z','AA']

	value=0
	while (value<(len(fusion_report_column_list))):
		fusion_report_column=fusion_report_column_list[value]
		summary_column=summary_column_list[value]



		ws5[summary_column+'23']= '=IF(OR(Gene_fusion_report!Y25<>"Artefact",Gene_fusion_report!Z25<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'25, "")'
		ws5[summary_column+'24']= '=IF(OR(Gene_fusion_report!Y26<>"Artefact",Gene_fusion_report!Z26<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'26, "")'
		ws5[summary_column+'25']= '=IF(OR(Gene_fusion_report!Y27<>"Artefact",Gene_fusion_report!Z27<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'27, "")'
		ws5[summary_column+'26']= '=IF(OR(Gene_fusion_report!Y28<>"Artefact",Gene_fusion_report!Z28<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'28, "")'
		ws5[summary_column+'27']= '=IF(OR(Gene_fusion_report!Y29<>"Artefact",Gene_fusion_report!Z29<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'29, "")'
		ws5[summary_column+'28']= '=IF(OR(Gene_fusion_report!Y30<>"Artefact",Gene_fusion_report!Z30<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'30, "")'
		ws5[summary_column+'29']= '=IF(OR(Gene_fusion_report!Y31<>"Artefact",Gene_fusion_report!Z31<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'31, "")'
		ws5[summary_column+'30']= '=IF(OR(Gene_fusion_report!Y32<>"Artefact",Gene_fusion_report!Z32<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'32, "")'
		ws5[summary_column+'31']= '=IF(OR(Gene_fusion_report!Y33<>"Artefact",Gene_fusion_report!Z33<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'33, "")'
		ws5[summary_column+'32']= '=IF(OR(Gene_fusion_report!Y34<>"Artefact",Gene_fusion_report!Z34<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'34, "")'
		ws5[summary_column+'33']= '=IF(OR(Gene_fusion_report!Y35<>"Artefact",Gene_fusion_report!Z35<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'35, "")'
		ws5[summary_column+'34']= '=IF(OR(Gene_fusion_report!Y36<>"Artefact",Gene_fusion_report!Z36<>"Artefact"),Gene_fusion_report!'+fusion_report_column+'36, "")'

		value=value+1


	font_bold=Font(bold=True)
	position= ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1']
	for cell in position:
		ws7[cell].font=font_bold


	colour= PatternFill("solid", fgColor="DCDCDC")
	position= ['A1','D1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1']
	for cell in position:
		ws7[cell].fill=colour

	colour= PatternFill("solid", fgColor="00FFFF00")
	position= ['B1', 'C1', 'E1', 'B2', 'C2', 'E2', 'O1', 'O2']
	for cell in position:
		ws7[cell].fill=colour


	colour= PatternFill("solid", fgColor="00FFCC00")
	position= ['A6', 'A7']
	for cell in position:
		ws7[cell].fill=colour


	border_a=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
	position=['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1', 'A6', 'A7', 'B6', 'B7']
	for cell in position:
		ws7[cell].border=border_a


	ws7['B2']=sampleId
	ws7['E2']=referral
	ws7['M2']=worksheetId
	ws7['N2']=seqId


	ws7.column_dimensions['A'].width=30
	ws7.column_dimensions['B'].width=20
	ws7.column_dimensions['C'].width=30
	ws7.column_dimensions['N'].width=40


	#format the summary spreadsheet by wrapping text, middle aligning and aligning left
	for row in ws5.iter_rows(min_row=3, max_row=4):
			for cell in row:
				cell.alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')


	for row in ws5.iter_rows(min_row=8, max_row=20):
			for cell in row:
				cell.alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')

		
	for row in ws5.iter_rows(min_row=37, max_row=39):
		for cell in row:
			cell.alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')


	for row in ws5.iter_rows(min_row=22, max_row=33):
		for cell in row:
			cell.alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')

		
	for row in ws5.iter_rows(min_row=42, max_row=45):
		for cell in row:
			cell.alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')


	# headers don't need their their text wrapped so only apply middle aligning and aligning left
	rows_list=[1,6,7,21,36,41,46]
	for value in rows_list:
		row=ws5[value]
		for cell in row:
			cell.alignment=Alignment(horizontal='left', vertical='center')

	return (wb, ws1, ws2, ws3, ws4,ws5,ws6,ws7,ws9)


def get_alignment_metrics_rmdup(path,sampleId, wb, ws5):

	#get the percentage aligned reads rmdup and number of aligned reads rmdup
	with open (path+sampleId+"_AlignmentSummaryMetrics_rmdup.txt") as file_rmdup:
		for line in file_rmdup:
			if line.startswith("CATEGORY"):
				headers_rmdup=line.split('\t')
			if line.startswith("PAIR"):
				pair_list_rmdup=line.split('\t')


	alignment_metrics_rmdup=pandas.DataFrame([pair_list_rmdup], columns=headers_rmdup)
	total_reads_rmdup=alignment_metrics_rmdup[['PF_READS']]
	total_reads_aligned_rmdup=alignment_metrics_rmdup[['PF_HQ_ALIGNED_READS']]
	total_reads_value_rmdup=total_reads_rmdup.iloc[0,0]
	aligned_reads_value_rmdup=total_reads_aligned_rmdup.iloc[0,0]
	percentage_aligned_reads_rmdup=(int(aligned_reads_value_rmdup)/int(total_reads_value_rmdup))*100
	ws5["G4"]=percentage_aligned_reads_rmdup
	aligned_reads_value_rmdup=int(aligned_reads_value_rmdup)
	ws5["B39"]=aligned_reads_value_rmdup

	#colour the cell red if number of aligned reads rmdup is less than the threshold of 38000
	if (aligned_reads_value_rmdup<38000):
		colour= PatternFill("solid", fgColor="00FF0000")
		position= ['B45']
		for cell in position:
			ws5[cell].fill=colour

	return (aligned_reads_value_rmdup, percentage_aligned_reads_rmdup, wb, ws5)
		 



def get_subpanel_NTC_check(coverage, aligned_reads_value_rmdup, path, sampleId, ntc, referral_list, coverage_rmdup, wb, ws9):


	ws9.column_dimensions['A'].width=40

	ws9["A13"]="NTC check for Arriba/STAR-Fusion referral genes" 
	ws9["B14"]=sampleId
	ws9["C14"]="NTC"
	ws9["D14"]="%NTC contamination"
	ws9["E14"]="Difference"
	ws9["A15"]="Total mapped reads(duplicates removed)" 
	ws9["B15"]= aligned_reads_value_rmdup


	#get the number of aligned reads without duplicates in the ntc and calculate the %ntc contamination
	with open (path+ntc+"_AlignmentSummaryMetrics_rmdup.txt") as file_rmdup:
		for line in file_rmdup:
			if line.startswith("CATEGORY"):
				headers_rmdup=line.split('\t')
			if line.startswith("PAIR"):
				pair_list_rmdup=line.split('\t')


	alignment_metrics_rmdup_ntc=pandas.DataFrame([pair_list_rmdup], columns=headers_rmdup)
	total_reads_aligned_rmdup_ntc=alignment_metrics_rmdup_ntc[['PF_HQ_ALIGNED_READS']]
	aligned_reads_value_rmdup_ntc=total_reads_aligned_rmdup_ntc.iloc[0,0]
	aligned_reads_value_rmdup_ntc=int(aligned_reads_value_rmdup_ntc)
	ws9["C15"]=aligned_reads_value_rmdup_ntc
	ws9["D15"]=((aligned_reads_value_rmdup_ntc/aligned_reads_value_rmdup)*100)
	ws9["E15"]='=B15-C15'

	ws9["A16"]= "Post PCR1 Qubit"
	ws9["B16"]='=Patient_demographics!O2'
	ws9["D16"]='=(C16/B16)*100'
	ws9["E16"]='=B16-C16'


	ws9["A21"]="NTC check for RMATS events(METex14skp and EGFRv3)" 
	ws9["A22"]="without duplicates" 
	ws9["A23"]="Feature"
	ws9["B23"]="AVG_DEPTH"
	ws9["C23"]="NTC_AVG_DEPTH"
	ws9["D23"]="%NTC contamination"

	font_bold=Font(bold=True)
	position= ['A13','A21']
	for cell in position:
		ws9[cell].font=font_bold


	#Add the EGFR and MET hotspot coverage values without duplicates removed to the subpanel NTC check tab

	if (("MET_exon14_skipping" in referral_list) or ("EGFRv3" in referral_list)):


		if ("MET_exon14_skipping" in referral_list):

			MET_exon_13_rmdup=coverage_rmdup[coverage_rmdup["START"]==116411698]
			MET_exon_15_rmdup=coverage_rmdup[coverage_rmdup["END"]==116414944]
			MET_exon_13_rmdup_ws9=MET_exon_13_rmdup.drop(["CHR", "START","END"], axis=1)
			MET_exon_15_rmdup_ws9=MET_exon_15_rmdup.drop(["CHR", "START","END"], axis=1)

			for row in dataframe_to_rows(MET_exon_13_rmdup_ws9, header=False, index=False):
				ws9.append(row)
			for row in dataframe_to_rows(MET_exon_15_rmdup_ws9, header=False, index=False):
				ws9.append(row)


		if  ("EGFRv3" in referral_list):

			EGFR_exon_1_rmdup=coverage_rmdup[coverage_rmdup["START"]==55087048]
			EGFR_exon_8_rmdup=coverage_rmdup[coverage_rmdup["END"]==55223532]
			EGFR_exon_1_rmdup_ws9=EGFR_exon_1_rmdup.drop(["CHR", "START","END"], axis=1)
			EGFR_exon_8_rmdup_ws9=EGFR_exon_8_rmdup.drop(["CHR", "START","END"], axis=1)


			for row in dataframe_to_rows(EGFR_exon_1_rmdup_ws9, header=False, index=False):
				ws9.append(row)
			for row in dataframe_to_rows(EGFR_exon_8_rmdup_ws9, header=False, index=False):
				ws9.append(row)


	#Add and format the NTC checker boxes in the subpanel NTC tab
	ws9['H7']="NTC check 1"
	ws9['H8']="NTC check 2"
	ws9['I6']="Results"
	ws9['J6']="Comments"

	border_a=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
	position=['I6','J6','H7','I7','J7','H8', 'I8', 'J8']
	for cell in position:
		ws9[cell].border=border_a

	colour= PatternFill("solid", fgColor="00CCFFFF")
	position= ['I6','J6']
	for cell in position:
		ws9[cell].fill=colour

	colour= PatternFill("solid", fgColor="DCDCDC")
	position= ['H7','H8']
	for cell in position:
		ws9[cell].fill=colour

	#add a PASS/FAIL dropdown box to the NTC CHECK
	dv = DataValidation(type="list", formula1='"PASS,FAIL,PARTIAL_FAIL"', allow_blank=True)
	ws9.add_data_validation(dv)
	dv.add(ws9['I7'])
	dv.add(ws9['I8'])


	#colour the %NTC contamination tab if the contamination level is above 10%
	colour= PatternFill("solid", start_color="00FF0000", end_color="00FF0000")
	ws9.conditional_formatting.add('D4:D9', CellIsRule( operator='greaterThan', formula=['10'], stopIfTrue=True, fill=colour))
	ws9.conditional_formatting.add('D15:D21', CellIsRule( operator='greaterThan', formula=['10'], stopIfTrue=True, fill=colour))


	return(ws9, wb)


def get_met_exon_skipping(referral_list, sampleId, referral, worksheetId, seqId, path,ntc, coverage_rmdup, wb, ws5, ws6):

	ws8=wb.create_sheet("RMATS")

	ws8['A4']='Lab number'
	ws8['B4']='Patient Name'
	ws8['C4']='Reason for referral'
	ws8['D4']='NGS Wks'
	ws8['E4']='Nextseq run ID'
	ws8['F4']='Checker 1 initials and date'
	ws8['G4']='Checker 2 initials and date'

	ws8['A8']='RMATS output'

	ws8['C1']='Analysis Sheet- Pan RNA Cancer Fusion Panel'

	ws8['A5']=sampleId
	ws8['B5']='=Patient_demographics!C2'
	ws8['C5']=referral
	ws8['D5']=worksheetId
	ws8['E5']=seqId

	ws8["A9"]="GeneID"
	ws8["B9"]="geneSymbol"
	ws8["C9"]="chr"
	ws8["D9"]="exonStart_0base"
	ws8["E9"]="exonEnd"
	ws8["F9"]="IJC_SAMPLE_1"
	ws8["G9"]="SJC_SAMPLE_1"
	ws8["H9"]="IJC_SAMPLE_2"
	ws8["I9"]="SJC_SAMPLE_2"
	ws8["J9"]="FDR"
	ws8["K9"]="IncLevel1"
	ws8["L9"]="IncLevel2"
	ws8["M9"]="IncLevelDifference"
	ws8["N9"]="Conclusion checker 1"
	ws8["O9"]="Conclusion checker 2"
	ws8["P9"]="Comments"

	ws5["A41"]="Intragenic fusions(RMATS) genuine calls"
	ws5['A46']="Intragenic fusions quality metrics"
	ws5["A42"]="geneSymbol"
	ws5["B42"]="chr"
	ws5["C42"]="exonStart_0base"
	ws5["D42"]="exonEnd"
	ws5["E42"]="IJC_SAMPLE_1"
	ws5["F42"]="SJC_SAMPLE_1"
	ws5["G42"]="FDR"
	ws5["H42"]="IncLevelDifference"
	ws5["O42"]="Conclusion Checker 1"
	ws5["P42"]="Conclusion Checker 2"
	ws5["Q42"]="Comments"


	ws5["A47"]="CHR"
	ws5["B47"]="START"
	ws5["C47"]="END"
	ws5["D47"]="AVG_DEPTH"
	ws5["E47"]="Gene_exon"
	ws5["F47"]="Conclusion 1"
	ws5["G47"]="Conclusion 2"



	colour= PatternFill("solid", fgColor="DCDCDC")
	position= ['A42', 'B42', 'C42', 'D42', 'E42', 'F42', 'G42', 'H42', 'I42', 'J42', 'K42', 'L42', 'M42', 'N42', 
                   'A47', 'B47', 'C47', 'D47', 'E47' ]
	for cell in position:
		ws5[cell].fill=colour

	colour= PatternFill("solid", fgColor="00FFCC00")
	position= ['O42', 'P42', 'Q42', 'F47', 'G47']
	for cell in position:
		ws5[cell].fill=colour

	
	border_a=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
	position=['A42','B42','C42','D42','E42','F42','G42', 'H42', 'I42', 'J42', 'K42', 'L42', 'M42', 'N42', 'O42', 'P42', 'Q42',
        'A47','B47','C47','D47','E47','F47','G47']
	for cell in position:
		ws5[cell].border=border_a



	#add the rmats report to the RMATS report tab. Only add exon skipping events for genes that are in the referral type.
	rmats=[]
	headers=[]
	line_number=0

	with open (path+sampleId+"/"+seqId+"_"+sampleId+"_RMATS_Report.tsv") as file:
		for line in file:
			if line.startswith("#"):
				continue
			else:
				if (line_number==0):
					headers=line.split('\t')
					line_number=1
				else:	
					rmats_line=line.split('\t')
					rmats.append(rmats_line)
	rmats_dataframe = pandas.DataFrame(rmats, columns=headers)
	rmats_dataframe=rmats_dataframe[['GeneID', 'geneSymbol', 'chr', 'exonStart_0base', 'exonEnd', 'IJC_SAMPLE_1', 'SJC_SAMPLE_1', 'IJC_SAMPLE_2', 'SJC_SAMPLE_2', 'FDR', 'IncLevel1', 'IncLevel2', 'IncLevelDifference']]



	for gene in referral_list:
		if (gene=="MET_exon14_skipping"):
			rmats_MET=rmats_dataframe[rmats_dataframe["geneSymbol"]=='"MET"']
			if (len(rmats_MET)>0):
				print("MET_exon_14_skipping present")
			else:
				dict_RMATS=({"GeneID":"", "geneSymbol": "MET_exon_14_skipping-no fusions found", "chr": "", "exonStart_0base": "", "exonEnd": "", "IJC_SAMPLE_1":"", "SJC_SAMPLE_1": "", "IJC_SAMPLE_2": "", "SJC_SAMPLE_2":"", "FDR": "", "IncLevel1": "", "IncLevel2": "", "IncLevelDifference": ""}) 
				rmats_MET= pandas.DataFrame(dict_RMATS, columns=["GeneID", "geneSymbol", "chr", "exonStart_0base", "exonEnd", "IJC_SAMPLE_1", "SJC_SAMPLE_1", "IJC_SAMPLE_2","SJC_SAMPLE_2" "FDR", "IncLevel1", "IncLevel2", "IncLevelDifference"], index=[0]) 

			for row in dataframe_to_rows(rmats_MET, header=False, index=False):
				ws8.append(row)


		if (gene=="EGFRv3"):
			rmats_EGFR=rmats_dataframe[rmats_dataframe["geneSymbol"]=='"EGFR"']
			if (len(rmats_EGFR)>0):
				print("EGFRv3 present")
			
			else:
				dict_RMATS=({"GeneID":"", "geneSymbol": "EGFRv3-no fusions found", "chr": "", "exonStart_0base": "", "exonEnd": "", "IJC_SAMPLE_1":"", "SJC_SAMPLE_1": "", "IJC_SAMPLE_2": "", "SJC_SAMPLE_2":"", "FDR": "", "IncLevel1": "", "IncLevel2": "", "IncLevelDifference": ""}) 
				rmats_EGFR= pandas.DataFrame(dict_RMATS, columns=["GeneID", "geneSymbol", "chr", "exonStart_0base", "exonEnd", "IJC_SAMPLE_1", "SJC_SAMPLE_1", "IJC_SAMPLE_2","SJC_SAMPLE_2" "FDR", "IncLevel1", "IncLevel2", "IncLevelDifference"], index=[0]) 

			for row in dataframe_to_rows(rmats_EGFR, header=False, index=False):
				ws8.append(row)



	#add dropdown box in the conclusion columns of the RMATS report
	dv = DataValidation(type="list", formula1='"Genuine,Artefact,Fail_QC,no_fusions"', allow_blank=True)
	ws8.add_data_validation(dv)
	dv.add('N10:N12')
	dv.add('O10:O12')


	#transfer values in RMATS table in RMATS tab to the summary tab
	summary_column_list=['A', 'B','C','D','E','F','G','H','O','P', 'Q']
	RMATS_column_list=[ 'B','C','D','E','F','G','J','M','N', 'O', 'P']
	value=0
	while (value<(len(RMATS_column_list))):
		RMATS_column=RMATS_column_list[value]
		summary_column=summary_column_list[value]

		ws5[summary_column+'43']= '=IF(OR(RMATS!N10<>"Artefact",RMATS!O10<>"Artefact"),RMATS!'+RMATS_column+'10, "0")'
		ws5[summary_column+'44']= '=IF(OR(RMATS!N11<>"Artefact",RMATS!O11<>"Artefact"),RMATS!'+RMATS_column+'11, "0")'
		ws5[summary_column+'45']= '=IF(OR(RMATS!N12<>"Artefact",RMATS!O12<>"Artefact"),RMATS!'+RMATS_column+'12, "0")'
		value=value+1

	

	ws8.column_dimensions['A'].width=30
	ws8.column_dimensions['B'].width=15
	ws8.column_dimensions['C'].width=15
	ws8.column_dimensions['H'].width=70
	ws8.column_dimensions['I'].width=40
	ws8.column_dimensions['L'].width=40

	for column in ['D','E', 'F','G','J', 'k','M', 'N', 'O', 'P','Q']:
		ws8.column_dimensions[column].width=20


	font_bold=Font(bold=True)
	position= ['A4','B4','C4','D4','E4','F4','G4','A8','B8','C8','D8','E8','F8','G8','H8','I8','J8','K8','L8','M8','N8','O8','P8','Q8']
	for cell in position:
		ws8[cell].font=font_bold


	border_a=Border(left=Side(border_style=BORDER_MEDIUM), right=Side(border_style=BORDER_MEDIUM), top=Side(border_style=BORDER_MEDIUM), bottom=Side(border_style=BORDER_MEDIUM))
	position=['A4','B4','C4','D4','E4','F4','G4','A5','B5','C5','D5','E5','F5','G5','A9','B9','C9','D9','E9','F9','G9','H9','I9','J9','K9','L9','M9','N9','O9','P9']
	for cell in position:
		ws8[cell].border=border_a

	colour= PatternFill("solid", fgColor="DCDCDC")
	position= ['A9','B9','C9','D9','E9','F9','G9','H9','I9','J9','K9','L9','M9','N9','O9','P9']
	for cell in position:
		ws8[cell].fill=colour


	colour= PatternFill("solid", fgColor="00CCFFFF")
	position= ['A4','B4','C4','D4','E4','F4', 'G4']
	for cell in position:
		ws8[cell].fill=colour

	colour= PatternFill("solid", fgColor="00FFCC00")
	position= ['N9', 'O9', 'P9']
	for cell in position:
		ws8[cell].fill=colour

	
	ws6['A33']= "RMATS results"
	ws6["A34"]="GeneID"
	ws6["B34"]="geneSymbol"
	ws6["C34"]="chr"
	ws6["D34"]="exonStart_0base"
	ws6["E34"]="exonEnd"
	ws6["F34"]="IJC_SAMPLE_1"
	ws6["G34"]="SJC_SAMPLE_1"
	ws6["H34"]="IJC_SAMPLE_2"
	ws6["I34"]="SJC_SAMPLE_2"
	ws6["J34"]="FDR"
	ws6["K34"]="IncLevel1"
	ws6["L34"]="IncLevel2"
	ws6["M34"]="IncLevelDifference"
	ws6["N34"]="Conclusion checker 1"
	ws6["O34"]="Conclusion checker 2"
	ws6["P34"]="Comments"

	#add the rmats ntc report to the ntc report tab. Only add exon skipping events for genes that are in the referral type.
	rmats_ntc=[]
	headers=[]
	line_number=0

	with open (path+ntc+"/"+seqId+"_"+ntc+"_RMATS_Report.tsv") as file:
		for line in file:
			if line.startswith("#"):
				continue
			else:
				if (line_number==0):
					headers=line.split('\t')
					line_number=1
				else:	
					rmats_ntc_line=line.split('\t')
					rmats_ntc.append(rmats_ntc_line)
	rmats_ntc_dataframe = pandas.DataFrame(rmats_ntc, columns=headers)
	rmats_ntc_dataframe=rmats_ntc_dataframe[['GeneID', 'geneSymbol', 'chr', 'exonStart_0base', 'exonEnd', 'IJC_SAMPLE_1', 'SJC_SAMPLE_1', 'IJC_SAMPLE_2', 'SJC_SAMPLE_2', 'FDR', 'IncLevel1', 'IncLevel2', 'IncLevelDifference']]

	for gene in referral_list:
		if (gene=="MET_exon14_skipping"):
			rmats_ntc_MET=rmats_ntc_dataframe[rmats_ntc_dataframe["geneSymbol"]=='"MET"']
			if (len(rmats_ntc_MET)>0):
				print("MET_exon_14_skipping present in ntc")
			else:
				dict_RMATS_ntc=({"GeneID":"", "geneSymbol": "MET_exon_14_skipping-no fusions found", "chr": "", "exonStart_0base": "", "exonEnd": "", "IJC_SAMPLE_1":"", "SJC_SAMPLE_1": "", "IJC_SAMPLE_2": "", "SJC_SAMPLE_2":"", "FDR": "", "IncLevel1": "", "IncLevel2": "", "IncLevelDifference": ""}) 
				rmats_ntc_MET= pandas.DataFrame(dict_RMATS_ntc, columns=["GeneID", "geneSymbol", "chr", "exonStart_0base", "exonEnd", "IJC_SAMPLE_1", "SJC_SAMPLE_1", "IJC_SAMPLE_2","SJC_SAMPLE_2" "FDR", "IncLevel1", "IncLevel2", "IncLevelDifference"], index=[0]) 

			for row in dataframe_to_rows(rmats_ntc_MET, header=False, index=False):
				ws6.append(row)


		if (gene=="EGFRv3"):
			rmats_ntc_EGFR=rmats_ntc_dataframe[rmats_ntc_dataframe["geneSymbol"]=='"EGFR"']
			if (len(rmats_ntc_EGFR)>0):
				print("EGFRv3 present in ntc")
			
			else:
				dict_RMATS_ntc=({"GeneID":"", "geneSymbol": "EGFRv3-no fusions found", "chr": "", "exonStart_0base": "", "exonEnd": "", "IJC_SAMPLE_1":"", "SJC_SAMPLE_1": "", "IJC_SAMPLE_2": "", "SJC_SAMPLE_2":"", "FDR": "", "IncLevel1": "", "IncLevel2": "", "IncLevelDifference": ""}) 
				rmats_ntc_EGFR= pandas.DataFrame(dict_RMATS_ntc, columns=["GeneID", "geneSymbol", "chr", "exonStart_0base", "exonEnd", "IJC_SAMPLE_1", "SJC_SAMPLE_1", "IJC_SAMPLE_2","SJC_SAMPLE_2" "FDR", "IncLevel1", "IncLevel2", "IncLevelDifference"], index=[0]) 

			for row in dataframe_to_rows(rmats_ntc_EGFR, header=False, index=False):
				ws6.append(row)





	#Add the coverage values for EGFRv3 and MET_exon14_skipping events breakpoints into the summary tab and if the coverage is over the thresholds highlight in red
	MET_exon_13_rmdup=coverage_rmdup[coverage_rmdup["START"]==116411698]
	MET_exon_15_rmdup=coverage_rmdup[coverage_rmdup["END"]==116414944]
	EGFR_exon_1_rmdup=coverage_rmdup[coverage_rmdup["START"]==55087048]
	EGFR_exon_8_rmdup=coverage_rmdup[coverage_rmdup["END"]==55223532]

	del [MET_exon_13_rmdup["META"], MET_exon_13_rmdup["NTC_AVG_DEPTH"],MET_exon_13_rmdup["%NTC contamination"]]
	del [MET_exon_15_rmdup["META"],MET_exon_15_rmdup["NTC_AVG_DEPTH"],MET_exon_15_rmdup["%NTC contamination"]]
	del [EGFR_exon_1_rmdup["META"], EGFR_exon_1_rmdup["NTC_AVG_DEPTH"],EGFR_exon_1_rmdup["%NTC contamination"]]
	del [EGFR_exon_8_rmdup["META"], EGFR_exon_8_rmdup["NTC_AVG_DEPTH"], EGFR_exon_8_rmdup["%NTC contamination"]]


	#only add coverage values if "MET_exon14_skipping" is in the referral list
	if ("MET_exon14_skipping" in referral_list):
		ws5["A48"]="MET coverage results" 
		for row in dataframe_to_rows(MET_exon_13_rmdup, header=False, index=False):
			ws5.append(row)
		MET_exon_13_rmdup_AVGdepth=MET_exon_13_rmdup[["AVG_DEPTH"]]
		if (MET_exon_13_rmdup_AVGdepth.iloc[0,0]<7):
			colour= PatternFill("solid", fgColor="00FF0000")
			position= ['D49']
			for cell in position:
				ws5[cell].fill=colour


		for row in dataframe_to_rows(MET_exon_15_rmdup, header=False, index=False):
			ws5.append(row)
		MET_exon_15_rmdup_AVGdepth=MET_exon_15_rmdup[["AVG_DEPTH"]]
		if (MET_exon_15_rmdup_AVGdepth.iloc[0,0]<6):
			colour= PatternFill("solid", fgColor="00FF0000")
			position= ['D50']
			for cell in position:
				ws5[cell].fill=colour

		ws5["E49"]="MET_Exon13_last10bases"
		ws5["E50"]="MET_Exon15_first10bases"


	#only add coverage values if "EGFRv3" is in the referral list
	if ("EGFRv3" in referral_list):
		ws5["A51"]="EGFR coverage results"
		for row in dataframe_to_rows(EGFR_exon_1_rmdup, header=False, index=False):
			ws5.append(row)
		EGFR_exon_1_rmdup_AVGdepth=EGFR_exon_1_rmdup[["AVG_DEPTH"]]
		if (EGFR_exon_1_rmdup_AVGdepth.iloc[0,0]<6):
			colour= PatternFill("solid", fgColor="00FF0000")
			position= ['D52']
			for cell in position:
				ws5[cell].fill=colour


		for row in dataframe_to_rows(EGFR_exon_8_rmdup, header=False, index=False):
			ws5.append(row)
		EGFR_exon_8_rmdup_AVGdepth=EGFR_exon_8_rmdup[["AVG_DEPTH"]]
		if (EGFR_exon_8_rmdup_AVGdepth.iloc[0,0]<9):
			colour= PatternFill("solid", fgColor="00FF0000")
			position= ['D53']
			for cell in position:
				ws5[cell].fill=colour

		ws5["E52"]="EGFR_Exon1_last10bases"
		ws5["E53"]="EGFR_Exon8_first10bases"


	
	#format the RMATS part of the summary spreadsheet by wrapping text, middle aligning and aligning left

	for row in ws5.iter_rows(min_row=47, max_row=55):
		for cell in row:
			cell.alignment=Alignment(wrap_text=True, horizontal='left', vertical='center')


	return (ws5, ws8, wb, ws6)


if __name__ == "__main__":

	seqId=sys.argv[1]
	sampleId=sys.argv[2]
	referral=sys.argv[3]
	ntc=sys.argv[4]
	worksheetId=sys.argv[5]
	version=sys.argv[6]




	path="/data/results/"+seqId + "/RocheSTFusion/"

	referrals_path="/data/diagnostics/pipelines/SomaticFusion/SomaticFusion-"+version+"/RocheSTFusion/"





	wb=Workbook()
	ws7=wb.create_sheet("Patient_demographics")
	ws6= wb.create_sheet("NTC fusion report")
	ws9=wb.create_sheet("Subpanel_NTC_Check")
	ws1= wb.create_sheet("Gene_fusion_report")
	ws5=wb.create_sheet("Summary")
	ws3=wb.create_sheet("coverage_with_duplicates")
	ws4=wb.create_sheet("coverage_without_duplicates")
	ws2=wb.create_sheet("total_coverage")

	ws1["A4"]="Lab number"
	ws1["B4"]="Patient name"
	ws1["C4"]="Reason for referral"
	ws1["D4"]="NGS wks"
	ws1["E4"]="NextSeq runId"
	ws1["F4"]="Checker 1 initials and date"
	ws1["G4"]="Checker 2 initials and date"


	ws7['A1']='Date Received'
	ws7['B1']='LABNO'
	ws7['C1']='Patient name'
	ws7['D1']='DOB'
	ws7['E1']='Reason for referral'
	ws7['F1']='Referring Clinician'
	ws7['G1']='Indication'
	ws7['H1']='Due Date'
	ws7['I1']='% Tumour'
	ws7['J1']='DV200'
	ws7['K1']='Pre-processing (dil/zymo)'
	ws7['L1']='Qubit RNA conc after pre-processing'
	ws7['M1']='NGS Worksheet'
	ws7['N1']='NextSeq run ID'
	ws7['O1']='PostPCR1 Qubit'
	ws7['P1']='Result'
	ws7['Q1']='Date reported'
	ws7['R1']='Comments'

	ws7['A6']="Checker 1 initials and date"
	ws7['A7']="Checker 2 initals and date"

	ws5['H5'].number_format='mm-dd-yy'

	referral_list, wb=get_referral_list(referral, referrals_path, wb)
	ntc_star_fusion_report, ntc_arriba_report,wb, ws6=get_NTC_fusion_report(ntc, referral_list, path, wb, ws6)
	star_fusion_report_final,wb,ws1,ws5=get_star_fusion_report(referral_list, ntc_star_fusion_report, sampleId, path,wb, ws1,ws5)
	arriba_report,wb, ws1, ws5=get_arriba_fusion_report(referral_list, sampleId, path,wb, ws1, ws5)
	ntc_total_average_depth, ntc_total_average_depth_rmdup=get_ntc_total_coverage(ntc,path)
	total_coverage, total_coverage_rmdup, ws2,wb=get_total_coverage(ntc_total_average_depth, ntc_total_average_depth_rmdup,sampleId, path,wb, ws2)
	ntc_average_depth, ntc_average_depth_rmdup=ntc_get_coverage(path,ntc)
	coverage_rmdup, coverage, ws3, ws4, wb=get_coverage(referral, ntc_average_depth, ntc_average_depth_rmdup, sampleId,path, wb, ws3, ws4)
	wb, ws1, ws2, ws3,ws4,ws5,ws6,ws7, ws9=format_analysis_sheet(referral, wb, ws1, ws2, ws3,ws4, ws5, ws6, ws7,ws9)
	aligned_reads_value_rmdup, percentage_aligned_reads_rmdup,wb, ws5=get_alignment_metrics_rmdup(path,sampleId,wb,ws5)
	ws9,wb=get_subpanel_NTC_check(coverage, aligned_reads_value_rmdup, path, sampleId,ntc, referral_list,coverage_rmdup,wb,ws9)

	if (("MET_exon14_skipping" in referral_list) or ("EGFRv3" in referral_list)):
		ws5, ws8, wb,ws6=get_met_exon_skipping(referral_list, sampleId, referral, worksheetId, seqId, path, ntc, coverage_rmdup,wb, ws5, ws6)
	wb.save(path +sampleId+"_"+referral +".xlsx")
