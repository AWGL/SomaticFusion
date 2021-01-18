import unittest

from make_worksheets import *
import openpyxl

path="./tests/"
referrals_path="./RocheSTFusion/"


wb_Colorectal=Workbook()
ws7_Colorectal=wb_Colorectal.create_sheet("Patient_demographics")
ws6_Colorectal= wb_Colorectal.create_sheet("NTC fusion report")
ws9_Colorectal=wb_Colorectal.create_sheet("Subpanel_NTC_Check")
ws1_Colorectal= wb_Colorectal.create_sheet("Gene_fusion_report")
ws5_Colorectal=wb_Colorectal.create_sheet("Summary")
ws3_Colorectal=wb_Colorectal.create_sheet("coverage_with_duplicates")
ws4_Colorectal=wb_Colorectal.create_sheet("coverage_without_duplicates")
ws2_Colorectal=wb_Colorectal.create_sheet("total_coverage")

wb_GIST=Workbook()
ws7_GIST=wb_GIST.create_sheet("Patient_demographics")
ws6_GIST= wb_GIST.create_sheet("NTC fusion report")
ws9_GIST=wb_GIST.create_sheet("Subpanel_NTC_Check")
ws1_GIST= wb_GIST.create_sheet("Gene_fusion_report")
ws5_GIST=wb_GIST.create_sheet("Summary")
ws3_GIST=wb_GIST.create_sheet("coverage_with_duplicates")
ws4_GIST=wb_GIST.create_sheet("coverage_without_duplicates")
ws2_GIST=wb_GIST.create_sheet("total_coverage")


wb_Glioma=Workbook()
ws7_Glioma=wb_Glioma.create_sheet("Patient_demographics")
ws6_Glioma= wb_Glioma.create_sheet("NTC fusion report")
ws9_Glioma=wb_Glioma.create_sheet("Subpanel_NTC_Check")
ws1_Glioma= wb_Glioma.create_sheet("Gene_fusion_report")
ws5_Glioma=wb_Glioma.create_sheet("Summary")
ws3_Glioma=wb_Glioma.create_sheet("coverage_with_duplicates")	
ws4_Glioma=wb_Glioma.create_sheet("coverage_without_duplicates")
ws2_Glioma=wb_Glioma.create_sheet("total_coverage")

wb_Lung=Workbook()
ws7_Lung=wb_Lung.create_sheet("Patient_demographics")
ws6_Lung= wb_Lung.create_sheet("NTC fusion report")
ws9_Lung=wb_Lung.create_sheet("Subpanel_NTC_Check")
ws1_Lung= wb_Lung.create_sheet("Gene_fusion_report")
ws5_Lung=wb_Lung.create_sheet("Summary")
ws3_Lung=wb_Lung.create_sheet("coverage_with_duplicates")	
ws4_Lung=wb_Lung.create_sheet("coverage_without_duplicates")
ws2_Lung=wb_Lung.create_sheet("total_coverage")

wb_Melanoma=Workbook()
ws7_Melanoma=wb_Melanoma.create_sheet("Patient_demographics")
ws6_Melanoma= wb_Melanoma.create_sheet("NTC fusion report")
ws9_Melanoma=wb_Melanoma.create_sheet("Subpanel_NTC_Check")
ws1_Melanoma= wb_Melanoma.create_sheet("Gene_fusion_report")
ws5_Melanoma=wb_Melanoma.create_sheet("Summary")
ws3_Melanoma=wb_Melanoma.create_sheet("coverage_with_duplicates")
ws4_Melanoma=wb_Melanoma.create_sheet("coverage_without_duplicates")
ws2_Melanoma=wb_Melanoma.create_sheet("total_coverage")


wb_NTRK=Workbook()
ws7_NTRK=wb_NTRK.create_sheet("Patient_demographics")
ws6_NTRK= wb_NTRK.create_sheet("NTC fusion report")
ws9_NTRK=wb_NTRK.create_sheet("Subpanel_NTC_Check")
ws1_NTRK= wb_NTRK.create_sheet("Gene_fusion_report")
ws5_NTRK=wb_NTRK.create_sheet("Summary")
ws3_NTRK=wb_NTRK.create_sheet("coverage_with_duplicates")
ws4_NTRK=wb_NTRK.create_sheet("coverage_without_duplicates")
ws2_NTRK=wb_NTRK.create_sheet("total_coverage")



wb_Thyroid=Workbook()
ws7_Thyroid=wb_Thyroid.create_sheet("Patient_demographics")
ws6_Thyroid= wb_Thyroid.create_sheet("NTC fusion report")
ws9_Thyroid=wb_Thyroid.create_sheet("Subpanel_NTC_Check")
ws1_Thyroid= wb_Thyroid.create_sheet("Gene_fusion_report")
ws5_Thyroid=wb_Thyroid.create_sheet("Summary")
ws3_Thyroid=wb_Thyroid.create_sheet("coverage_with_duplicates")
ws4_Thyroid=wb_Thyroid.create_sheet("coverage_without_duplicates")
ws2_Thyroid=wb_Thyroid.create_sheet("total_coverage")


wb_Tumour=Workbook()
ws7_Tumour=wb_Tumour.create_sheet("Patient_demographics")
ws6_Tumour= wb_Tumour.create_sheet("NTC fusion report")
ws9_Tumour=wb_Tumour.create_sheet("Subpanel_NTC_Check")
ws1_Tumour= wb_Tumour.create_sheet("Gene_fusion_report")
ws5_Tumour=wb_Tumour.create_sheet("Summary")
ws3_Tumour=wb_Tumour.create_sheet("coverage_with_duplicates")
ws4_Tumour=wb_Tumour.create_sheet("coverage_without_duplicates")
ws2_Tumour=wb_Tumour.create_sheet("total_coverage")

referral_list_Colorectal=["NTRK1","NTRK2","NTRK3"]
referral_list_GIST=["NTRK1","NTRK2","NTRK3"]
referral_list_Glioma=["EGFRv3", "BRAF","NTRK1", "NTRK2","NTRK3"]
referral_list_Lung=["EGFRv3","ALK","MET_exon14_skipping","RET","ROS1","NTRK1","NTRK2","NTRK3"]
referral_list_Melanoma=["NTRK1","NTRK2","NTRK3"]
referral_list_NTRK=["NTRK1","NTRK2","NTRK3"]
referral_list_Thyroid=["EGFRv3","RET","NTRK1","NTRK2", "NTRK3"]
referral_list_Tumour=["ALK", "BRAF", "EGFRv3", "MET_exon14_skipping", "NTRK1", "NTRK2", "NTRK3", "RET", "ROS1"]

NTC_star_fusion_report_Colorectal=""
NTC_star_fusion_report_GIST=""
NTC_star_fusion_report_Glioma=""
NTC_star_fusion_report_Lung=""
NTC_star_fusion_report_Melanoma=""
NTC_star_fusion_report_NTRK=""
NTC_star_fusion_report_Thyroid=""
NTC_star_fusion_report_Tumour=""



class test_make_worksheets(unittest.TestCase):

	def test_get_referral_list(self):

		wb1=Workbook()
		ws7=wb1.create_sheet("Patient_demographics")
		ws6= wb1.create_sheet("NTC fusion report")
		ws9= wb1.create_sheet("Subpanel_NTC_Check")
		ws1= wb1.create_sheet("Gene_fusion_report")
		ws5= wb1.create_sheet("Summary")
		ws3= wb1.create_sheet("coverage_with_duplicates")
		ws4= wb1.create_sheet("coverage_without_duplicates")
		ws2= wb1.create_sheet("total_coverage")

		self.assertEqual(len((get_referral_list("Colorectal", referrals_path,wb1))[0]),3)
		self.assertEqual(len((get_referral_list("GIST", referrals_path,wb1)[0])),3)
		self.assertEqual(len((get_referral_list("Glioma", referrals_path,wb1))[0]),5)
		self.assertEqual(len((get_referral_list("Lung", referrals_path,wb1))[0]),8)
		self.assertEqual(len((get_referral_list("Melanoma", referrals_path,wb1))[0]),3)
		self.assertEqual(len((get_referral_list("NTRK", referrals_path,wb1)[0])),3)
		self.assertEqual(len((get_referral_list("Thyroid", referrals_path,wb1))[0]),5)
		self.assertEqual(len((get_referral_list("Tumour", referrals_path,wb1))[0]),9)


	def test_get_NTC_fusion_report(self):

		#Colorectal
		NTC_star_fusion_report_Colorectal, NTC_arriba_report_Colorectal,wb_Colorectal_output,ws6_Colorectal_output=(get_NTC_fusion_report("test-NTC",referral_list_Colorectal, path, wb_Colorectal, ws6_Colorectal))
		self.assertEqual(ws6_Colorectal["A5"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Colorectal["A6"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Colorectal["A7"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Colorectal["A8"].value,None)
		self.assertEqual(ws6_Colorectal["A9"].value,None)
		self.assertEqual(ws6_Colorectal["A10"].value,None)
		self.assertEqual(ws6_Colorectal["A11"].value,None)
		self.assertEqual(ws6_Colorectal["A12"].value, None)

		self.assertEqual(ws6_Colorectal["A20"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Colorectal["A21"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Colorectal["A22"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Colorectal["A23"].value,None)
		self.assertEqual(ws6_Colorectal["A24"].value,None)
		self.assertEqual(ws6_Colorectal["A25"].value,None)
		self.assertEqual(ws6_Colorectal["A26"].value,None)
		self.assertEqual(ws6_Colorectal["A27"].value, None)

		#GIST
		NTC_star_fusion_report_GIST, NTC_arriba_report_GIST,wb_GIST_output,ws6_GIST_output=(get_NTC_fusion_report("test-NTC",referral_list_GIST, path, wb_GIST, ws6_GIST))
		self.assertEqual(ws6_GIST["A5"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_GIST["A6"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_GIST["A7"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_GIST["A8"].value,None)
		self.assertEqual(ws6_GIST["A9"].value,None)
		self.assertEqual(ws6_GIST["A10"].value,None)
		self.assertEqual(ws6_GIST["A11"].value,None)
		self.assertEqual(ws6_GIST["A12"].value, None)

		self.assertEqual(ws6_GIST["A20"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_GIST["A21"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_GIST["A22"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_GIST["A23"].value,None)
		self.assertEqual(ws6_GIST["A24"].value,None)
		self.assertEqual(ws6_GIST["A25"].value,None)
		self.assertEqual(ws6_GIST["A26"].value,None)
		self.assertEqual(ws6_GIST["A27"].value, None)

		#Glioma
		NTC_star_fusion_report_Glioma, NTC_arriba_report_Glioma,wb_Glioma_output,ws6_Glioma_output=(get_NTC_fusion_report("test-NTC",referral_list_Glioma, path, wb_Glioma, ws6_Glioma))
		self.assertEqual(ws6_Glioma["A5"].value,"BRAF-no fusions found")
		self.assertEqual(ws6_Glioma["A6"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Glioma["A7"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Glioma["A8"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Glioma["A9"].value,None)
		self.assertEqual(ws6_Glioma["A10"].value,None)
		self.assertEqual(ws6_Glioma["A11"].value,None)
		self.assertEqual(ws6_Glioma["A12"].value, None)

		self.assertEqual(ws6_Glioma["A20"].value,"BRAF-no fusions found")
		self.assertEqual(ws6_Glioma["A21"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Glioma["A22"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Glioma["A23"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Glioma["A24"].value,None)
		self.assertEqual(ws6_Glioma["A25"].value,None)
		self.assertEqual(ws6_Glioma["A26"].value,None)
		self.assertEqual(ws6_Glioma["A27"].value, None)

		#Lung
		NTC_star_fusion_report_Lung, NTC_arriba_report_Lung,wb_Lung_output,ws6_Lung_output=(get_NTC_fusion_report("test-NTC",referral_list_Lung, path, wb_Lung, ws6_Lung))
		self.assertEqual(ws6_Lung["A5"].value,"ALK-no fusions found")
		self.assertEqual(ws6_Lung["A6"].value,"RET-no fusions found")
		self.assertEqual(ws6_Lung["A7"].value,"ROS1-no fusions found")
		self.assertEqual(ws6_Lung["A8"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Lung["A9"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Lung["A10"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Lung["A11"].value,None)
		self.assertEqual(ws6_Lung["A12"].value, None)

		self.assertEqual(ws6_Lung["A20"].value,"ALK-no fusions found")
		self.assertEqual(ws6_Lung["A21"].value,"RET-no fusions found")
		self.assertEqual(ws6_Lung["A22"].value,"ROS1-no fusions found")
		self.assertEqual(ws6_Lung["A23"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Lung["A24"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Lung["A25"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Lung["A26"].value,None)
		self.assertEqual(ws6_Lung["A27"].value, None)

		#Melanoma
		NTC_star_fusion_report_Melanoma, NTC_arriba_report_Melanoma,wb_Melanoma_output,ws6_Melanoma_output=(get_NTC_fusion_report("test-NTC",referral_list_Melanoma, path, wb_Melanoma, ws6_Melanoma))
		self.assertEqual(ws6_Melanoma["A5"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Melanoma["A6"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Melanoma["A7"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Melanoma["A8"].value,None)
		self.assertEqual(ws6_Melanoma["A9"].value,None)
		self.assertEqual(ws6_Melanoma["A10"].value,None)
		self.assertEqual(ws6_Melanoma["A11"].value,None)
		self.assertEqual(ws6_Melanoma["A12"].value, None)

		self.assertEqual(ws6_Melanoma["A20"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Melanoma["A21"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Melanoma["A22"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Melanoma["A23"].value,None)
		self.assertEqual(ws6_Melanoma["A24"].value,None)
		self.assertEqual(ws6_Melanoma["A25"].value,None)
		self.assertEqual(ws6_Melanoma["A26"].value,None)
		self.assertEqual(ws6_Melanoma["A27"].value, None)


		#NTRK
		NTC_star_fusion_report_NTRK, NTC_arriba_report_NTRK,wb_NTRK_output,ws6_NTRK_output=(get_NTC_fusion_report("test-NTC",referral_list_NTRK, path, wb_NTRK, ws6_NTRK))
		self.assertEqual(ws6_NTRK["A5"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_NTRK["A6"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_NTRK["A7"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_NTRK["A8"].value,None)
		self.assertEqual(ws6_NTRK["A9"].value,None)
		self.assertEqual(ws6_NTRK["A10"].value,None)
		self.assertEqual(ws6_NTRK["A11"].value,None)
		self.assertEqual(ws6_NTRK["A12"].value, None)

		self.assertEqual(ws6_NTRK["A20"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_NTRK["A21"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_NTRK["A22"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_NTRK["A23"].value,None)
		self.assertEqual(ws6_NTRK["A24"].value,None)
		self.assertEqual(ws6_NTRK["A25"].value,None)
		self.assertEqual(ws6_NTRK["A26"].value,None)
		self.assertEqual(ws6_NTRK["A27"].value, None)

		#Thyroid
		NTC_star_fusion_report_Thyroid, NTC_arriba_report_Thyroid,wb_Thyroid_output,ws6_Thyroid_output=(get_NTC_fusion_report("test-NTC",referral_list_Thyroid, path, wb_Thyroid, ws6_Thyroid))
		self.assertEqual(ws6_Thyroid["A5"].value,"RET-no fusions found")
		self.assertEqual(ws6_Thyroid["A6"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Thyroid["A7"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Thyroid["A8"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Thyroid["A9"].value,None)
		self.assertEqual(ws6_Thyroid["A10"].value,None)
		self.assertEqual(ws6_Thyroid["A11"].value,None)
		self.assertEqual(ws6_Thyroid["A12"].value, None)

		self.assertEqual(ws6_Thyroid["A20"].value,"RET-no fusions found")
		self.assertEqual(ws6_Thyroid["A21"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Thyroid["A22"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Thyroid["A23"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Thyroid["A24"].value,None)
		self.assertEqual(ws6_Thyroid["A25"].value,None)
		self.assertEqual(ws6_Thyroid["A26"].value,None)
		self.assertEqual(ws6_Thyroid["A27"].value, None)

		#Tumour
		NTC_star_fusion_report_Tumour, NTC_arriba_report_Tumour,wb_Tumour_output,ws6_Tumour_output=(get_NTC_fusion_report("test-NTC",referral_list_Tumour, path, wb_Tumour, ws6_Tumour))
		self.assertEqual(ws6_Tumour["A5"].value,"ALK-no fusions found")
		self.assertEqual(ws6_Tumour["A6"].value,"BRAF-no fusions found")
		self.assertEqual(ws6_Tumour["A7"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Tumour["A8"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Tumour["A9"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Tumour["A10"].value,"RET-no fusions found")
		self.assertEqual(ws6_Tumour["A11"].value,"ROS1-no fusions found")
		self.assertEqual(ws6_Tumour["A12"].value, None)

		self.assertEqual(ws6_Tumour["A20"].value,"ALK-no fusions found")
		self.assertEqual(ws6_Tumour["A21"].value,"BRAF-no fusions found")
		self.assertEqual(ws6_Tumour["A22"].value,"NTRK1-no fusions found")
		self.assertEqual(ws6_Tumour["A23"].value,"NTRK2-no fusions found")
		self.assertEqual(ws6_Tumour["A24"].value,"NTRK3-no fusions found")
		self.assertEqual(ws6_Tumour["A25"].value,"RET-no fusions found")
		self.assertEqual(ws6_Tumour["A26"].value,"ROS1-no fusions found")
		self.assertEqual(ws6_Tumour["A27"].value, None)



	def test_get_star_fusion_report(self):
		star_fusion_report_final_Colorectal, wb_Colorectal_output, ws1_Colorectal_output, ws5_Colorectal_output= get_star_fusion_report(referral_list_Colorectal, NTC_star_fusion_report_Colorectal, "test-sample", path,wb_Colorectal, ws1_Colorectal, ws5_Colorectal)
		self.assertEqual(ws1_Colorectal_output["A11"].value, "NTRK1--Gene7")
		self.assertEqual(ws1_Colorectal_output["A12"].value, "NTRK1--Gene10")		
		self.assertEqual(ws1_Colorectal_output["A13"].value, "NTRK2--Gene3")
		self.assertEqual(ws1_Colorectal_output["A14"].value, "Gene4--NTRK2")
		self.assertEqual(ws1_Colorectal_output["A15"].value, "NTRK3--Gene3")
		self.assertEqual(ws1_Colorectal_output["A16"].value, "Gene4--NTRK3")
		self.assertEqual(ws1_Colorectal_output["A17"].value, "Gene5--NTRK3")		
		self.assertEqual(ws1_Colorectal_output["A18"].value, "Gene6--NTRK3")
		self.assertEqual(ws1_Colorectal_output["A19"].value, "NTRK3--Gene7")
		self.assertEqual(ws1_Colorectal_output["A20"].value, None)
		self.assertEqual(ws1_Colorectal_output["A21"].value, None)
		self.assertEqual(ws1_Colorectal_output["A22"].value, None)
		self.assertEqual(ws1_Colorectal_output["A23"].value, None)
		self.assertEqual(ws1_Colorectal_output["A24"].value, None)

		star_fusion_report_final_GIST, wb_GIST_output, ws1_GIST_output, ws5_GIST_output= get_star_fusion_report(referral_list_GIST, NTC_star_fusion_report_GIST, "test-sample", path,wb_GIST, ws1_GIST, ws5_GIST)
		self.assertEqual(ws1_GIST_output["A11"].value, "NTRK1--Gene7")
		self.assertEqual(ws1_GIST_output["A12"].value, "NTRK1--Gene10")		
		self.assertEqual(ws1_GIST_output["A13"].value, "NTRK2--Gene3")
		self.assertEqual(ws1_GIST_output["A14"].value, "Gene4--NTRK2")
		self.assertEqual(ws1_GIST_output["A15"].value, "NTRK3--Gene3")
		self.assertEqual(ws1_GIST_output["A16"].value, "Gene4--NTRK3")
		self.assertEqual(ws1_GIST_output["A17"].value, "Gene5--NTRK3")		
		self.assertEqual(ws1_GIST_output["A18"].value, "Gene6--NTRK3")
		self.assertEqual(ws1_GIST_output["A19"].value, "NTRK3--Gene7")
		self.assertEqual(ws1_GIST_output["A20"].value, None)
		self.assertEqual(ws1_GIST_output["A21"].value, None)
		self.assertEqual(ws1_GIST_output["A22"].value, None)
		self.assertEqual(ws1_GIST_output["A23"].value, None)
		self.assertEqual(ws1_GIST_output["A24"].value, None)

		star_fusion_report_final_Glioma, wb_Glioma_output, ws1_Glioma_output, ws5_Glioma_output= get_star_fusion_report(referral_list_Glioma, NTC_star_fusion_report_Glioma, "test-sample", path,wb_Glioma, ws1_Glioma, ws5_Glioma)
		self.assertEqual(ws1_Glioma_output["A11"].value, "BRAF-no fusions found")
		self.assertEqual(ws1_Glioma_output["A12"].value, "NTRK1--Gene7")
		self.assertEqual(ws1_Glioma_output["A13"].value, "NTRK1--Gene10")		
		self.assertEqual(ws1_Glioma_output["A14"].value, "NTRK2--Gene3")
		self.assertEqual(ws1_Glioma_output["A15"].value, "Gene4--NTRK2")
		self.assertEqual(ws1_Glioma_output["A16"].value, "NTRK3--Gene3")
		self.assertEqual(ws1_Glioma_output["A17"].value, "Gene4--NTRK3")
		self.assertEqual(ws1_Glioma_output["A18"].value, "Gene5--NTRK3")		
		self.assertEqual(ws1_Glioma_output["A19"].value, "Gene6--NTRK3")
		self.assertEqual(ws1_Glioma_output["A20"].value, "NTRK3--Gene7")
		self.assertEqual(ws1_Glioma_output["A21"].value, None)
		self.assertEqual(ws1_Glioma_output["A22"].value, None)
		self.assertEqual(ws1_Glioma_output["A23"].value, None)
		self.assertEqual(ws1_Glioma_output["A24"].value, None)
		self.assertEqual(ws1_Glioma_output["A25"].value, None)

		star_fusion_report_final_Lung, wb_Lung_output, ws1_Lung_output, ws5_Lung_output= get_star_fusion_report(referral_list_Lung, NTC_star_fusion_report_Lung, "test-sample", path,wb_Lung, ws1_Lung, ws5_Lung)
		self.assertEqual(ws1_Lung_output["A11"].value, "ALK--Gene3")
		self.assertEqual(ws1_Lung_output["A12"].value, "Gene5--ALK")		
		self.assertEqual(ws1_Lung_output["A13"].value, "Gene2--RET")
		self.assertEqual(ws1_Lung_output["A14"].value, "RET--Gene3")
		self.assertEqual(ws1_Lung_output["A15"].value, "ROS1--Gene1")
		self.assertEqual(ws1_Lung_output["A16"].value, "NTRK1--Gene7")
		self.assertEqual(ws1_Lung_output["A17"].value, "NTRK1--Gene10")		
		self.assertEqual(ws1_Lung_output["A18"].value, "NTRK2--Gene3")
		self.assertEqual(ws1_Lung_output["A19"].value, "Gene4--NTRK2")
		self.assertEqual(ws1_Lung_output["A20"].value, "NTRK3--Gene3")
		self.assertEqual(ws1_Lung_output["A21"].value, "Gene4--NTRK3")
		self.assertEqual(ws1_Lung_output["A22"].value, "ERROR:NOT ENOUGH SPACE FOR ALL FUSIONS-CHECK RESULTS FILES")
		self.assertEqual(ws1_Lung_output["A23"].value, "Gene6--NTRK3")
		self.assertEqual(ws1_Lung_output["A24"].value, "NTRK3--Gene7")
		self.assertEqual(ws1_Lung_output["A25"].value, None)

		star_fusion_report_final_Melanoma, wb_Melanoma_output, ws1_Melanoma_output, ws5_Melanoma_output= get_star_fusion_report(referral_list_Melanoma, NTC_star_fusion_report_Melanoma, "test-sample", path,wb_Melanoma, ws1_Melanoma, ws5_Melanoma)
		self.assertEqual(ws1_Melanoma_output["A11"].value, "NTRK1--Gene7")
		self.assertEqual(ws1_Melanoma_output["A12"].value, "NTRK1--Gene10")		
		self.assertEqual(ws1_Melanoma_output["A13"].value, "NTRK2--Gene3")
		self.assertEqual(ws1_Melanoma_output["A14"].value, "Gene4--NTRK2")
		self.assertEqual(ws1_Melanoma_output["A15"].value, "NTRK3--Gene3")
		self.assertEqual(ws1_Melanoma_output["A16"].value, "Gene4--NTRK3")
		self.assertEqual(ws1_Melanoma_output["A17"].value, "Gene5--NTRK3")		
		self.assertEqual(ws1_Melanoma_output["A18"].value, "Gene6--NTRK3")
		self.assertEqual(ws1_Melanoma_output["A19"].value, "NTRK3--Gene7")
		self.assertEqual(ws1_Melanoma_output["A20"].value, None)
		self.assertEqual(ws1_Melanoma_output["A21"].value, None)
		self.assertEqual(ws1_Melanoma_output["A22"].value, None)
		self.assertEqual(ws1_Melanoma_output["A23"].value, None)
		self.assertEqual(ws1_Melanoma_output["A24"].value, None)

		star_fusion_report_final_NTRK, wb_NTRK_output, ws1_NTRK_output, ws5_NTRK_output= get_star_fusion_report(referral_list_NTRK, NTC_star_fusion_report_NTRK, "test-sample", path,wb_NTRK, ws1_NTRK, ws5_NTRK)
		self.assertEqual(ws1_NTRK_output["A11"].value, "NTRK1--Gene7")
		self.assertEqual(ws1_NTRK_output["A12"].value, "NTRK1--Gene10")		
		self.assertEqual(ws1_NTRK_output["A13"].value, "NTRK2--Gene3")
		self.assertEqual(ws1_NTRK_output["A14"].value, "Gene4--NTRK2")
		self.assertEqual(ws1_NTRK_output["A15"].value, "NTRK3--Gene3")
		self.assertEqual(ws1_NTRK_output["A16"].value, "Gene4--NTRK3")
		self.assertEqual(ws1_NTRK_output["A17"].value, "Gene5--NTRK3")		
		self.assertEqual(ws1_NTRK_output["A18"].value, "Gene6--NTRK3")
		self.assertEqual(ws1_NTRK_output["A19"].value, "NTRK3--Gene7")
		self.assertEqual(ws1_NTRK_output["A20"].value, None)
		self.assertEqual(ws1_NTRK_output["A21"].value, None)
		self.assertEqual(ws1_NTRK_output["A22"].value, None)
		self.assertEqual(ws1_NTRK_output["A23"].value, None)
		self.assertEqual(ws1_NTRK_output["A24"].value, None)

		star_fusion_report_final_Thyroid, wb_Thyroid_output, ws1_Thyroid_output, ws5_Thyroid_output= get_star_fusion_report(referral_list_Thyroid, NTC_star_fusion_report_Thyroid, "test-sample", path,wb_Thyroid, ws1_Thyroid, ws5_Thyroid)
		self.assertEqual(ws1_Thyroid_output["A11"].value, "Gene2--RET")
		self.assertEqual(ws1_Thyroid_output["A12"].value, "RET--Gene3")		
		self.assertEqual(ws1_Thyroid_output["A13"].value, "NTRK1--Gene7")
		self.assertEqual(ws1_Thyroid_output["A14"].value, "NTRK1--Gene10")
		self.assertEqual(ws1_Thyroid_output["A15"].value, "NTRK2--Gene3")
		self.assertEqual(ws1_Thyroid_output["A16"].value, "Gene4--NTRK2")
		self.assertEqual(ws1_Thyroid_output["A17"].value, "NTRK3--Gene3")		
		self.assertEqual(ws1_Thyroid_output["A18"].value, "Gene4--NTRK3")
		self.assertEqual(ws1_Thyroid_output["A19"].value, "Gene5--NTRK3")
		self.assertEqual(ws1_Thyroid_output["A20"].value, "Gene6--NTRK3")
		self.assertEqual(ws1_Thyroid_output["A21"].value, "NTRK3--Gene7")
		self.assertEqual(ws1_Thyroid_output["A22"].value, None)
		self.assertEqual(ws1_Thyroid_output["A23"].value, None)
		self.assertEqual(ws1_Thyroid_output["A24"].value, None)
		self.assertEqual(ws1_Thyroid_output["A25"].value, None)

		star_fusion_report_final_Tumour, wb_Tumour_output, ws1_Tumour_output, ws5_Tumour_output= get_star_fusion_report(referral_list_Tumour, NTC_star_fusion_report_Tumour, "test-sample", path,wb_Tumour, ws1_Tumour, ws5_Tumour)
		self.assertEqual(ws1_Tumour_output["A11"].value, "ALK--Gene3")
		self.assertEqual(ws1_Tumour_output["A12"].value, "Gene5--ALK")		
		self.assertEqual(ws1_Tumour_output["A13"].value, "BRAF-no fusions found")
		self.assertEqual(ws1_Tumour_output["A14"].value, "NTRK1--Gene7")
		self.assertEqual(ws1_Tumour_output["A15"].value, "NTRK1--Gene10")
		self.assertEqual(ws1_Tumour_output["A16"].value, "NTRK2--Gene3")
		self.assertEqual(ws1_Tumour_output["A17"].value, "Gene4--NTRK2")		
		self.assertEqual(ws1_Tumour_output["A18"].value, "NTRK3--Gene3")
		self.assertEqual(ws1_Tumour_output["A19"].value, "Gene4--NTRK3")
		self.assertEqual(ws1_Tumour_output["A20"].value, "Gene5--NTRK3")
		self.assertEqual(ws1_Tumour_output["A21"].value, "Gene6--NTRK3")
		self.assertEqual(ws1_Tumour_output["A22"].value, "ERROR:NOT ENOUGH SPACE FOR ALL FUSIONS-CHECK RESULTS FILES")
		self.assertEqual(ws1_Tumour_output["A23"].value, "Gene2--RET")
		self.assertEqual(ws1_Tumour_output["A24"].value, "RET--Gene3")
		self.assertEqual(ws1_Tumour_output["A25"].value, "ROS1--Gene1")
		self.assertEqual(ws1_Tumour_output["A26"].value, None)


	def test_get_arriba_fusion_report(self):

		#Colorectal
		wb_Colorectal2=Workbook()
		ws7_Colorectal2=wb_Colorectal.create_sheet("Patient_demographics")
		ws6_Colorectal2= wb_Colorectal.create_sheet("NTC fusion report")
		ws9_Colorectal2=wb_Colorectal2.create_sheet("Subpanel_NTC_Check")
		ws1_Colorectal2= wb_Colorectal2.create_sheet("Gene_fusion_report")
		ws5_Colorectal2=wb_Colorectal2.create_sheet("Summary")
		ws3_Colorectal2=wb_Colorectal2.create_sheet("coverage_with_duplicates")
		ws4_Colorectal2=wb_Colorectal2.create_sheet("coverage_without_duplicates")
		ws2_Colorectal2=wb_Colorectal2.create_sheet("total_coverage")

		arriba_report_final_Colorectal, wb_Colorectal_output, ws1_Colorectal_output, ws5_Colorectal_output= get_arriba_fusion_report(referral_list_GIST, "test-sample", path,wb_Colorectal2, ws1_Colorectal2, ws5_Colorectal2)
		self.assertEqual(ws1_Colorectal_output["A25"].value, "gene1")
		self.assertEqual(ws1_Colorectal_output["A26"].value, "NTRK1")
		self.assertEqual(ws1_Colorectal_output["A27"].value, "NTRK2-no fusions found")
		self.assertEqual(ws1_Colorectal_output["A28"].value, "NTRK3-no fusions found")
		self.assertEqual(ws1_Colorectal_output["A29"].value, None)
		self.assertEqual(ws1_Colorectal_output["A30"].value, None)
		self.assertEqual(ws1_Colorectal_output["A31"].value, None)
		self.assertEqual(ws1_Colorectal_output["A32"].value, None)

		self.assertEqual(ws1_Colorectal_output["B25"].value, "NTRK1")
		self.assertEqual(ws1_Colorectal_output["B26"].value, "gene2")
		self.assertEqual(ws1_Colorectal_output["B27"].value, '')
		self.assertEqual(ws1_Colorectal_output["B28"].value, '')
		self.assertEqual(ws1_Colorectal_output["B29"].value, None)
		self.assertEqual(ws1_Colorectal_output["B30"].value, None)
		self.assertEqual(ws1_Colorectal_output["B31"].value, None)
		self.assertEqual(ws1_Colorectal_output["B32"].value, None)

		#Glioma
		wb_Glioma2=Workbook()
		ws7_Glioma=wb_Glioma2.create_sheet("Patient_demographics")
		ws6_Glioma2= wb_Glioma2.create_sheet("NTC fusion report")
		ws9_Glioma2=wb_Glioma2.create_sheet("Subpanel_NTC_Check")
		ws1_Glioma2= wb_Glioma2.create_sheet("Gene_fusion_report")
		ws5_Glioma2=wb_Glioma2.create_sheet("Summary")
		ws3_Glioma2=wb_Glioma2.create_sheet("coverage_with_duplicates")
		ws4_Glioma2=wb_Glioma2.create_sheet("coverage_without_duplicates")
		ws2_Glioma2=wb_Glioma2.create_sheet("total_coverage")

		arriba_report_final_Glioma, wb_Glioma_output, ws1_Glioma_output, ws5_Glioma_output= get_arriba_fusion_report(referral_list_Glioma, "test-sample", path,wb_Glioma2, ws1_Glioma2, ws5_Glioma2)
		self.assertEqual(ws1_Glioma_output["A25"].value, "BRAF")
		self.assertEqual(ws1_Glioma_output["A26"].value, "gene1")
		self.assertEqual(ws1_Glioma_output["A27"].value, "BRAF")
		self.assertEqual(ws1_Glioma_output["A28"].value, "gene1")
		self.assertEqual(ws1_Glioma_output["A29"].value, "BRAF")
		self.assertEqual(ws1_Glioma_output["A30"].value, "gene1")
		self.assertEqual(ws1_Glioma_output["A31"].value, "NTRK1")
		self.assertEqual(ws1_Glioma_output["A32"].value, "NTRK2-no fusions found")
		self.assertEqual(ws1_Glioma_output["A33"].value, "NTRK3-no fusions found")
		self.assertEqual(ws1_Glioma_output["A34"].value, None)

		self.assertEqual(ws1_Glioma_output["B25"].value, "gene2")
		self.assertEqual(ws1_Glioma_output["B26"].value, "BRAF")
		self.assertEqual(ws1_Glioma_output["B27"].value, "gene2")
		self.assertEqual(ws1_Glioma_output["B28"].value, "BRAF")
		self.assertEqual(ws1_Glioma_output["B29"].value, "gene2")
		self.assertEqual(ws1_Glioma_output["B30"].value, "NTRK1")
		self.assertEqual(ws1_Glioma_output["B31"].value, "gene2")
		self.assertEqual(ws1_Glioma_output["B32"].value, '')
		self.assertEqual(ws1_Glioma_output["B33"].value, '')
		self.assertEqual(ws1_Glioma_output["B34"].value, None)

		#GIST
		wb_GIST2=Workbook()
		ws7_GIST2=wb_GIST2.create_sheet("Patient_demographics")
		ws6_GIST2= wb_GIST2.create_sheet("NTC fusion report")
		ws9_GIST2=wb_GIST2.create_sheet("Subpanel_NTC_Check")
		ws1_GIST2= wb_GIST2.create_sheet("Gene_fusion_report")
		ws5_GIST2=wb_GIST2.create_sheet("Summary")
		ws3_GIST2=wb_GIST2.create_sheet("coverage_with_duplicates")
		ws4_GIST2=wb_GIST2.create_sheet("coverage_without_duplicates")
		ws2_GIST2=wb_GIST2.create_sheet("total_coverage")

		arriba_report_final_GIST, wb_GIST_output, ws1_GIST_output, ws5_GIST_output= get_arriba_fusion_report(referral_list_GIST, "test-sample", path,wb_GIST2, ws1_GIST2, ws5_GIST2)
		self.assertEqual(ws1_GIST_output["A25"].value, "gene1")
		self.assertEqual(ws1_GIST_output["A26"].value, "NTRK1")
		self.assertEqual(ws1_GIST_output["A27"].value, "NTRK2-no fusions found")
		self.assertEqual(ws1_GIST_output["A28"].value, "NTRK3-no fusions found")
		self.assertEqual(ws1_GIST_output["A29"].value, None)
		self.assertEqual(ws1_GIST_output["A30"].value, None)
		self.assertEqual(ws1_GIST_output["A31"].value, None)
		self.assertEqual(ws1_GIST_output["A32"].value, None)

		self.assertEqual(ws1_GIST_output["B25"].value, "NTRK1")
		self.assertEqual(ws1_GIST_output["B26"].value, "gene2")
		self.assertEqual(ws1_GIST_output["B27"].value, '')
		self.assertEqual(ws1_GIST_output["B28"].value, '')
		self.assertEqual(ws1_GIST_output["B29"].value, None)
		self.assertEqual(ws1_GIST_output["B30"].value, None)
		self.assertEqual(ws1_GIST_output["B31"].value, None)
		self.assertEqual(ws1_GIST_output["B32"].value, None)

		#Lung
		wb_Lung2=Workbook()
		ws7_Lung2=wb_Lung2.create_sheet("Patient_demographics")
		ws6_Lung2= wb_Lung2.create_sheet("NTC fusion report")
		ws9_Lung2=wb_Lung2.create_sheet("Subpanel_NTC_Check")
		ws1_Lung2= wb_Lung2.create_sheet("Gene_fusion_report")
		ws5_Lung2=wb_Lung2.create_sheet("Summary")
		ws3_Lung2=wb_Lung2.create_sheet("coverage_with_duplicates")
		ws4_Lung2=wb_Lung2.create_sheet("coverage_without_duplicates")
		ws2_Lung2=wb_Lung2.create_sheet("total_coverage")

		arriba_report_final_Lung, wb_Lung_output, ws1_Lung_output, ws5_Lung_output= get_arriba_fusion_report(referral_list_Lung, "test-sample", path,wb_Lung2, ws1_Lung2, ws5_Lung2)
		self.assertEqual(ws1_Lung_output["A25"].value, "gene1")
		self.assertEqual(ws1_Lung_output["A26"].value, "ALK")
		self.assertEqual(ws1_Lung_output["A27"].value, "gene2")
		self.assertEqual(ws1_Lung_output["A28"].value, "ALK")
		self.assertEqual(ws1_Lung_output["A29"].value, "gene3")
		self.assertEqual(ws1_Lung_output["A30"].value, "ALK")
		self.assertEqual(ws1_Lung_output["A31"].value, "RET-no fusions found")
		self.assertEqual(ws1_Lung_output["A32"].value, "ROS1-no fusions found")
		self.assertEqual(ws1_Lung_output["A33"].value, "gene1")
		self.assertEqual(ws1_Lung_output["A34"].value, "NTRK1")
		self.assertEqual(ws1_Lung_output["A35"].value, "NTRK2-no fusions found")
		self.assertEqual(ws1_Lung_output["A36"].value, "NTRK3-no fusions found")
		self.assertEqual(ws1_Lung_output["A37"].value, None)

		self.assertEqual(ws1_Lung_output["B25"].value, "ALK")
		self.assertEqual(ws1_Lung_output["B26"].value, "gene4")
		self.assertEqual(ws1_Lung_output["B27"].value,"ALK")
		self.assertEqual(ws1_Lung_output["B28"].value, "gene5")
		self.assertEqual(ws1_Lung_output["B29"].value, "ALK")
		self.assertEqual(ws1_Lung_output["B30"].value, "gene6")
		self.assertEqual(ws1_Lung_output["B31"].value, "")
		self.assertEqual(ws1_Lung_output["B32"].value, "")
		self.assertEqual(ws1_Lung_output["B33"].value, "NTRK1")
		self.assertEqual(ws1_Lung_output["B34"].value, "gene2")
		self.assertEqual(ws1_Lung_output["B35"].value, "")
		self.assertEqual(ws1_Lung_output["B36"].value, "")
		self.assertEqual(ws1_Lung_output["B37"].value, None)

		#Melanoma
		wb_Melanoma2=Workbook()
		ws7_Melanoma2=wb_Melanoma2.create_sheet("Patient_demographics")
		ws6_Melanoma2= wb_Melanoma2.create_sheet("NTC fusion report")
		ws9_Melanoma2=wb_Melanoma2.create_sheet("Subpanel_NTC_Check")
		ws1_Melanoma2= wb_Melanoma2.create_sheet("Gene_fusion_report")
		ws5_Melanoma2=wb_Melanoma2.create_sheet("Summary")
		ws3_Melanoma2=wb_Melanoma2.create_sheet("coverage_with_duplicates")
		ws4_Melanoma2=wb_Melanoma2.create_sheet("coverage_without_duplicates")
		ws2_Melanoma2=wb_Melanoma2.create_sheet("total_coverage")

		arriba_report_final_Melanoma, wb_Melanoma_output, ws1_Melanoma_output, ws5_Melanoma_output= get_arriba_fusion_report(referral_list_Melanoma, "test-sample", path,wb_Melanoma2, ws1_Melanoma2, ws5_Melanoma2)
		self.assertEqual(ws1_Melanoma_output["A25"].value, "gene1")
		self.assertEqual(ws1_Melanoma_output["A26"].value, "NTRK1")
		self.assertEqual(ws1_Melanoma_output["A27"].value, "NTRK2-no fusions found")
		self.assertEqual(ws1_Melanoma_output["A28"].value, "NTRK3-no fusions found")
		self.assertEqual(ws1_Melanoma_output["A29"].value, None)
		self.assertEqual(ws1_Melanoma_output["A30"].value, None)
		self.assertEqual(ws1_Melanoma_output["A31"].value, None)
		self.assertEqual(ws1_Melanoma_output["A32"].value, None)

		self.assertEqual(ws1_Melanoma_output["B25"].value, "NTRK1")
		self.assertEqual(ws1_Melanoma_output["B26"].value, "gene2")
		self.assertEqual(ws1_Melanoma_output["B27"].value, '')
		self.assertEqual(ws1_Melanoma_output["B28"].value, '')
		self.assertEqual(ws1_Melanoma_output["B29"].value, None)
		self.assertEqual(ws1_Melanoma_output["B30"].value, None)
		self.assertEqual(ws1_Melanoma_output["B31"].value, None)
		self.assertEqual(ws1_Melanoma_output["B32"].value, None)

		#NTRK
		wb_NTRK2=Workbook()
		ws7_NTRK2=wb_NTRK2.create_sheet("Patient_demographics")
		ws6_NTRK2= wb_NTRK2.create_sheet("NTC fusion report")
		ws9_NTRK2=wb_NTRK2.create_sheet("Subpanel_NTC_Check")
		ws1_NTRK2= wb_NTRK2.create_sheet("Gene_fusion_report")
		ws5_NTRK2=wb_NTRK2.create_sheet("Summary")
		ws3_NTRK2=wb_NTRK2.create_sheet("coverage_with_duplicates")
		ws4_NTRK2=wb_NTRK2.create_sheet("coverage_without_duplicates")
		ws2_NTRK2=wb_NTRK2.create_sheet("total_coverage")

		arriba_report_final_NTRK, wb_NTRK_output, ws1_NTRK_output, ws5_NTRK_output= get_arriba_fusion_report(referral_list_NTRK, "test-sample", path,wb_NTRK2, ws1_NTRK2, ws5_NTRK2)
		self.assertEqual(ws1_NTRK_output["A25"].value, "gene1")
		self.assertEqual(ws1_NTRK_output["A26"].value, "NTRK1")
		self.assertEqual(ws1_NTRK_output["A27"].value, "NTRK2-no fusions found")
		self.assertEqual(ws1_NTRK_output["A28"].value, "NTRK3-no fusions found")
		self.assertEqual(ws1_NTRK_output["A29"].value, None)
		self.assertEqual(ws1_NTRK_output["A30"].value, None)
		self.assertEqual(ws1_NTRK_output["A31"].value, None)
		self.assertEqual(ws1_NTRK_output["A32"].value, None)

		self.assertEqual(ws1_NTRK_output["B25"].value, "NTRK1")
		self.assertEqual(ws1_NTRK_output["B26"].value, "gene2")
		self.assertEqual(ws1_NTRK_output["B27"].value, '')
		self.assertEqual(ws1_NTRK_output["B28"].value, '')
		self.assertEqual(ws1_NTRK_output["B29"].value, None)
		self.assertEqual(ws1_NTRK_output["B30"].value, None)
		self.assertEqual(ws1_NTRK_output["B31"].value, None)
		self.assertEqual(ws1_NTRK_output["B32"].value, None)

		#Thyroid
		wb_Thyroid2=Workbook()
		ws7_Thyroid2=wb_Thyroid2.create_sheet("Patient_demographics")
		ws6_Thyroid2= wb_Thyroid2.create_sheet("NTC fusion report")
		ws9_Thyroid2=wb_Thyroid2.create_sheet("Subpanel_NTC_Check")
		ws1_Thyroid2= wb_Thyroid2.create_sheet("Gene_fusion_report")
		ws5_Thyroid2=wb_Thyroid2.create_sheet("Summary")
		ws3_Thyroid2=wb_Thyroid2.create_sheet("coverage_with_duplicates")
		ws4_Thyroid2=wb_Thyroid2.create_sheet("coverage_without_duplicates")
		ws2_Thyroid2=wb_Thyroid2.create_sheet("total_coverage")

		arriba_report_final_Thyroid, wb_Thyroid_output, ws1_Thyroid_output, ws5_Thyroid_output= get_arriba_fusion_report(referral_list_Thyroid, "test-sample", path,wb_Thyroid2, ws1_Thyroid2, ws5_Thyroid2)
		self.assertEqual(ws1_Thyroid_output["A25"].value, "RET-no fusions found")
		self.assertEqual(ws1_Thyroid_output["A26"].value, "gene1")
		self.assertEqual(ws1_Thyroid_output["A27"].value, "NTRK1")
		self.assertEqual(ws1_Thyroid_output["A28"].value, "NTRK2-no fusions found")
		self.assertEqual(ws1_Thyroid_output["A29"].value, "NTRK3-no fusions found")
		self.assertEqual(ws1_Thyroid_output["A30"].value, None)
		self.assertEqual(ws1_Thyroid_output["A31"].value, None)
		self.assertEqual(ws1_Thyroid_output["A32"].value, None)

		self.assertEqual(ws1_Thyroid_output["B25"].value, "")
		self.assertEqual(ws1_Thyroid_output["B26"].value, "NTRK1")
		self.assertEqual(ws1_Thyroid_output["B27"].value, "gene2")
		self.assertEqual(ws1_Thyroid_output["B28"].value, "")
		self.assertEqual(ws1_Thyroid_output["B29"].value, "")
		self.assertEqual(ws1_Thyroid_output["B30"].value, None)
		self.assertEqual(ws1_Thyroid_output["B31"].value, None)
		self.assertEqual(ws1_Thyroid_output["B32"].value, None)

		#Tumour
		wb_Tumour2=Workbook()
		ws7_Tumour2=wb_Tumour2.create_sheet("Patient_demographics")
		ws6_Tumour2= wb_Tumour2.create_sheet("NTC fusion report")
		ws9_Tumour2=wb_Tumour2.create_sheet("Subpanel_NTC_Check")
		ws1_Tumour2= wb_Tumour2.create_sheet("Gene_fusion_report")
		ws5_Tumour2=wb_Tumour2.create_sheet("Summary")
		ws3_Tumour2=wb_Tumour2.create_sheet("coverage_with_duplicates")
		ws4_Tumour2=wb_Tumour2.create_sheet("coverage_without_duplicates")
		ws2_Tumour2=wb_Tumour2.create_sheet("total_coverage")

		arriba_report_final_Tumour, wb_Tumour_output, ws1_Tumour_output, ws5_Tumour_output= get_arriba_fusion_report(referral_list_Tumour, "test-sample", path,wb_Tumour2, ws1_Tumour2, ws5_Tumour2)
		self.assertEqual(ws1_Tumour_output["A25"].value, "gene1")
		self.assertEqual(ws1_Tumour_output["A26"].value, "ALK")
		self.assertEqual(ws1_Tumour_output["A27"].value, "gene2")
		self.assertEqual(ws1_Tumour_output["A28"].value, "ALK")
		self.assertEqual(ws1_Tumour_output["A29"].value, "gene3")
		self.assertEqual(ws1_Tumour_output["A30"].value, "ALK")
		self.assertEqual(ws1_Tumour_output["A31"].value, "BRAF")
		self.assertEqual(ws1_Tumour_output["A32"].value, "gene1")
		self.assertEqual(ws1_Tumour_output["A33"].value, "BRAF")
		self.assertEqual(ws1_Tumour_output["A34"].value, "gene1")
		self.assertEqual(ws1_Tumour_output["A35"].value, "BRAF")
		self.assertEqual(ws1_Tumour_output["A36"].value, "ERROR:NOT ENOUGH SPACE FOR ALL FUSIONS-CHECK RESULTS FILES")
		self.assertEqual(ws1_Tumour_output["A37"].value, "NTRK1")
		self.assertEqual(ws1_Tumour_output["A38"].value, "NTRK2-no fusions found")
		self.assertEqual(ws1_Tumour_output["A39"].value, "NTRK3-no fusions found")
		self.assertEqual(ws1_Tumour_output["A40"].value, "RET-no fusions found")
		self.assertEqual(ws1_Tumour_output["A41"].value, "ROS1-no fusions found")
		self.assertEqual(ws1_Tumour_output["A42"].value, None)

		self.assertEqual(ws1_Tumour_output["B25"].value, "ALK")
		self.assertEqual(ws1_Tumour_output["B26"].value, "gene4")
		self.assertEqual(ws1_Tumour_output["B27"].value, "ALK")
		self.assertEqual(ws1_Tumour_output["B28"].value, "gene5")
		self.assertEqual(ws1_Tumour_output["B29"].value, "ALK")
		self.assertEqual(ws1_Tumour_output["B30"].value, "gene6")
		self.assertEqual(ws1_Tumour_output["B31"].value, "gene2")
		self.assertEqual(ws1_Tumour_output["B32"].value, "BRAF")
		self.assertEqual(ws1_Tumour_output["B33"].value, "gene2")
		self.assertEqual(ws1_Tumour_output["B34"].value, "BRAF")
		self.assertEqual(ws1_Tumour_output["B35"].value, "gene2")
		self.assertEqual(ws1_Tumour_output["B36"].value, "NTRK1")
		self.assertEqual(ws1_Tumour_output["B37"].value, "gene2")
		self.assertEqual(ws1_Tumour_output["B38"].value, "")
		self.assertEqual(ws1_Tumour_output["B39"].value, "")
		self.assertEqual(ws1_Tumour_output["B40"].value, "")
		self.assertEqual(ws1_Tumour_output["B41"].value, "")
		self.assertEqual(ws1_Tumour_output["B42"].value, None)


	def test_get_ntc_total_coverage(self):

		self.assertEqual(len(get_ntc_total_coverage("test-NTC", path)[0]),20)
		self.assertEqual(len(get_ntc_total_coverage("test-NTC", path)[1]),20)


	def test_get_total_coverage(self):

		ntc_total_coverage=pandas.read_csv(path+"test-NTC"+"/"+"test-NTC"+"_coverage.totalCoverage", sep="\t")
		ntc_total_average_depth=ntc_total_coverage["AVG_DEPTH"]

		ntc_total_coverage_rmdup=pandas.read_csv(path+"test-NTC"+"/"+"test-NTC"+"_rmdup_coverage.totalCoverage", sep="\t")
		ntc_total_average_depth_rmdup=ntc_total_coverage_rmdup["AVG_DEPTH"]

		wb=Workbook()
		ws7=wb.create_sheet("Patient_demographics")
		ws6= wb.create_sheet("NTC fusion report")
		ws9=wb.create_sheet("Subpanel_NTC_Check")
		ws1= wb.create_sheet("Gene_fusion_report")
		ws5=wb.create_sheet("Summary")
		ws3=wb.create_sheet("coverage_with_duplicates")
		ws4=wb.create_sheet("coverage_without_duplicates")
		ws2=wb.create_sheet("total_coverage")

		ws2=(get_total_coverage(ntc_total_average_depth, ntc_total_average_depth_rmdup, "test-sample", path,wb, ws2)[2])

		self.assertEqual(ws2["A4"].value,"Coverage_ROS1_GENE")
		self.assertEqual(ws2["B4"].value,57564)
		self.assertEqual(ws2["C4"].value,5.0)
		self.assertEqual(ws2["D4"].value,0.008685984295740393)

		self.assertEqual(ws2["A9"].value,"Coverage_NTRK3_GENE")
		self.assertEqual(ws2["B9"].value,75648)
		self.assertEqual(ws2["C9"].value,10)
		self.assertEqual(ws2["D9"].value,0.01321912013536379)

		self.assertEqual(ws2["A15"].value,"Coverage_EML4_GENE")
		self.assertEqual(ws2["B15"].value,96735)
		self.assertEqual(ws2["C15"].value,14)
		self.assertEqual(ws2["D15"].value,0.014472528040523078)

		self.assertEqual(ws2["A21"].value,"Coverage_SDC4_GENE")
		self.assertEqual(ws2["B21"].value,96749)
		self.assertEqual(ws2["C21"].value,7.0)
		self.assertEqual(ws2["D21"].value,0.007235216901466682)

		#total coverage rmdup
		self.assertEqual(ws2["A34"].value,"Coverage_ROS1_GENE")
		self.assertEqual(ws2["B34"].value,5957)
		self.assertEqual(ws2["C34"].value,57)
		self.assertEqual(ws2["D34"].value,0.9568574785966091)

		self.assertEqual(ws2["A39"].value,"Coverage_NTRK3_GENE")
		self.assertEqual(ws2["B39"].value,6584)
		self.assertEqual(ws2["C39"].value,16)
		self.assertEqual(ws2["D39"].value,0.24301336573511542)

		self.assertEqual(ws2["A45"].value,"Coverage_EML4_GENE")
		self.assertEqual(ws2["B45"].value,1432)
		self.assertEqual(ws2["C45"].value,5.0)
		self.assertEqual(ws2["D45"].value,0.34916201117318435)

		self.assertEqual(ws2["A51"].value,"Coverage_SDC4_GENE")
		self.assertEqual(ws2["B51"].value,3558)
		self.assertEqual(ws2["C51"].value,21)
		self.assertEqual(ws2["D51"].value,0.5902192242833052)


	def test_ntc_get_coverage(self):

		self.assertEqual(len(ntc_get_coverage(path, "test-NTC")[0]),286)
		self.assertEqual(len(ntc_get_coverage(path, "test-NTC")[1]),286)


	def test_get_coverage(self):

		ntc_total_coverage=pandas.read_csv(path+"test-NTC"+"/"+"test-NTC"+"_coverage.totalCoverage", sep="\t")
		ntc_total_average_depth=ntc_total_coverage["AVG_DEPTH"]

		ntc_total_coverage_rmdup=pandas.read_csv(path+"test-NTC"+"/"+"test-NTC"+"_rmdup_coverage.totalCoverage", sep="\t")
		ntc_total_average_depth_rmdup=ntc_total_coverage_rmdup["AVG_DEPTH"]

		ntc_average_depth=(ntc_get_coverage(path, "test-NTC"))[0]
		ntc_average_depth_rmdup=(ntc_get_coverage(path, "test-NTC"))[1]

		wb=Workbook()
		ws7=wb.create_sheet("Patient_demographics")
		ws6= wb.create_sheet("NTC fusion report")
		ws9=wb.create_sheet("Subpanel_NTC_Check")
		ws1= wb.create_sheet("Gene_fusion_report")
		ws5=wb.create_sheet("Summary")
		ws3=wb.create_sheet("coverage_with_duplicates")
		ws4=wb.create_sheet("coverage_without_duplicates")
		ws2=wb.create_sheet("total_coverage")

		#check the rmdup coverage file is correct

		coverage_rmdup, coverage, ws3, ws4,wb2=(get_coverage("Colorectal", ntc_average_depth, ntc_average_depth_rmdup, "test-sample", path,wb, ws3, ws4))

		self.assertEqual(ws4["B2"].value,117609654)
		self.assertEqual(ws4["C2"].value,117609965)
		self.assertEqual(ws4["D2"].value,"gene_name=ROS1;NM_002944.2;exon_num=43")
		self.assertEqual(ws4["E2"].value,52.0)
		self.assertEqual(ws4["F2"].value,9.0)
		self.assertEqual(ws4["G2"].value,17.307692307692307)

		self.assertEqual(ws4["B3"].value,117622136)
		self.assertEqual(ws4["C3"].value,117622300)
		self.assertEqual(ws4["D3"].value,"gene_name=ROS1;NM_002944.2;exon_num=42")
		self.assertEqual(ws4["E3"].value,77.0)
		self.assertEqual(ws4["F3"].value,11.0)
		self.assertEqual(ws4["G3"].value,14.285714285714285)

		self.assertEqual(ws4["B91"].value,87342568)
		self.assertEqual(ws4["C91"].value,87342874)
		self.assertEqual(ws4["D91"].value,"gene_name=NTRK2;NM_006180.4;exon_num=11")
		self.assertEqual(ws4["E91"].value,91.0)
		self.assertEqual(ws4["F91"].value,2.0)
		self.assertEqual(ws4["G91"].value,2.197802197802198)

		self.assertEqual(ws4["B106"].value,88476242)
		self.assertEqual(ws4["C106"].value,88476415)
		self.assertEqual(ws4["D106"].value,"gene_name=NTRK3;NM_001012338.2;exon_num=16")
		self.assertEqual(ws4["E106"].value,88.0)
		self.assertEqual(ws4["F106"].value,1.0)
		self.assertEqual(ws4["G106"].value,1.1363636363636365)

		self.assertEqual(ws4["B137"].value,116397691)
		self.assertEqual(ws4["C137"].value,116397828)
		self.assertEqual(ws4["D137"].value,"gene_name=MET;NM_001127500.1;exon_num=8")
		self.assertEqual(ws4["E137"].value,115.0)
		self.assertEqual(ws4["F137"].value,7.0)
		self.assertEqual(ws4["G137"].value,6.086956521739131)

		self.assertEqual(ws4["B156"].value,55220238)
		self.assertEqual(ws4["C156"].value,55220357)
		self.assertEqual(ws4["D156"].value,"gene_name=EGFR;NM_005228.3;exon_num=6")
		self.assertEqual(ws4["E156"].value,103.0)
		self.assertEqual(ws4["F156"].value,5.0)
		self.assertEqual(ws4["G156"].value,4.854368932038835)

		self.assertEqual(ws4["B188"].value,42491845)
		self.assertEqual(ws4["C188"].value,42491871)
		self.assertEqual(ws4["D188"].value,"gene_name=EML4;NM_019063.4;exon_num=6")
		self.assertEqual(ws4["E188"].value,46.0)
		self.assertEqual(ws4["F188"].value,2.0)
		self.assertEqual(ws4["G188"].value,4.3478260869565215)

		self.assertEqual(ws4["B214"].value,154163661)
		self.assertEqual(ws4["C214"].value,154163787)
		self.assertEqual(ws4["D214"].value,"gene_name=TPM3;NM_152263.2;exon_num=2")
		self.assertEqual(ws4["E214"].value,144.0)
		self.assertEqual(ws4["F214"].value,12.0)
		self.assertEqual(ws4["G214"].value,8.333333333333332)

		self.assertEqual(ws4["B255"].value,32327090)
		self.assertEqual(ws4["C255"].value,32327146)
		self.assertEqual(ws4["D255"].value,"gene_name=KIF5B;NM_004521.2;exon_num=6")
		self.assertEqual(ws4["E255"].value,83.0)
		self.assertEqual(ws4["F255"].value,10.0)
		self.assertEqual(ws4["G255"].value,12.048192771084338)

		self.assertEqual(ws4["B287"].value,159239113)
		self.assertEqual(ws4["C287"].value,159239125)
		self.assertEqual(ws4["D287"].value,"gene_name=EZR;NM_001111077.1;exon_num=2")
		self.assertEqual(ws4["E287"].value,91.0)
		self.assertEqual(ws4["F287"].value,8.0)
		self.assertEqual(ws4["G287"].value,8.791208791208792)

		#check the coverage file is correct

		self.assertEqual(ws3["B2"].value,117609654)
		self.assertEqual(ws3["C2"].value,117609965)
		self.assertEqual(ws3["D2"].value,"gene_name=ROS1;NM_002944.2;exon_num=43")
		self.assertEqual(ws3["E2"].value,5000)
		self.assertEqual(ws3["F2"].value,2.0)
		self.assertEqual(ws3["G2"].value,0.04)

		self.assertEqual(ws3["B3"].value,117622136)
		self.assertEqual(ws3["C3"].value,117622300)
		self.assertEqual(ws3["D3"].value,"gene_name=ROS1;NM_002944.2;exon_num=42")
		self.assertEqual(ws3["E3"].value,192)
		self.assertEqual(ws3["F3"].value,1.0)
		self.assertEqual(ws3["G3"].value,0.5208333333333333)

		self.assertEqual(ws3["B91"].value,87342568)
		self.assertEqual(ws3["C91"].value,87342874)
		self.assertEqual(ws3["D91"].value,"gene_name=NTRK2;NM_006180.4;exon_num=11")
		self.assertEqual(ws3["E91"].value,672)
		self.assertEqual(ws3["F91"].value,3.0)
		self.assertEqual(ws3["G91"].value,0.4464285714285714)

		self.assertEqual(ws3["B106"].value,88476242)
		self.assertEqual(ws3["C106"].value,88476415)
		self.assertEqual(ws3["D106"].value,"gene_name=NTRK3;NM_001012338.2;exon_num=16")
		self.assertEqual(ws3["E106"].value,1111)
		self.assertEqual(ws3["F106"].value,5.0)
		self.assertEqual(ws3["G106"].value,0.45004500450045004)

		self.assertEqual(ws3["B137"].value,116397691)
		self.assertEqual(ws3["C137"].value,116397828)
		self.assertEqual(ws3["D137"].value,"gene_name=MET;NM_001127500.1;exon_num=8")
		self.assertEqual(ws3["E137"].value,52.1)
		self.assertEqual(ws3["F137"].value,4.0)
		self.assertEqual(ws3["G137"].value,7.677543186180421)

		self.assertEqual(ws3["B156"].value,55220238)
		self.assertEqual(ws3["C156"].value,55220357)
		self.assertEqual(ws3["D156"].value,"gene_name=EGFR;NM_005228.3;exon_num=6")
		self.assertEqual(ws3["E156"].value,9341)
		self.assertEqual(ws3["F156"].value,1.0)
		self.assertEqual(ws3["G156"].value,0.010705491917353602)

		self.assertEqual(ws3["B188"].value,42491845)
		self.assertEqual(ws3["C188"].value,42491871)
		self.assertEqual(ws3["D188"].value,"gene_name=EML4;NM_019063.4;exon_num=6")
		self.assertEqual(ws3["E188"].value,731)
		self.assertEqual(ws3["F188"].value,6.0)
		self.assertEqual(ws3["G188"].value,0.8207934336525308)

		self.assertEqual(ws3["B214"].value,154163661)
		self.assertEqual(ws3["C214"].value,154163787)
		self.assertEqual(ws3["D214"].value,"gene_name=TPM3;NM_152263.2;exon_num=2")
		self.assertEqual(ws3["E214"].value,1468)
		self.assertEqual(ws3["F214"].value,8.0)
		self.assertEqual(ws3["G214"].value,0.544959128065395)

		self.assertEqual(ws3["B255"].value,32327090)
		self.assertEqual(ws3["C255"].value,32327146)
		self.assertEqual(ws3["D255"].value,"gene_name=KIF5B;NM_004521.2;exon_num=6")
		self.assertEqual(ws3["E255"].value,11)
		self.assertEqual(ws3["F255"].value,2.0)
		self.assertEqual(ws3["G255"].value,18.181818181818183)

		self.assertEqual(ws3["B287"].value,159239113)
		self.assertEqual(ws3["C287"].value,159239125)
		self.assertEqual(ws3["D287"].value,"gene_name=EZR;NM_001111077.1;exon_num=2")
		self.assertEqual(ws3["E287"].value,332)
		self.assertEqual(ws3["F287"].value,7.0)
		self.assertEqual(ws3["G287"].value,2.108433734939759)






	def test_get_alignment_metrics_rmdup(self):

		wb=Workbook()
		ws7=wb.create_sheet("Patient_demographics")
		ws6= wb.create_sheet("NTC fusion report")
		ws9=wb.create_sheet("Subpanel_NTC_Check")
		ws1= wb.create_sheet("Gene_fusion_report")
		ws5=wb.create_sheet("Summary")
		ws3=wb.create_sheet("coverage_with_duplicates")
		ws4=wb.create_sheet("coverage_without_duplicates")
		ws2=wb.create_sheet("total_coverage")

		self.assertEqual(get_alignment_metrics_rmdup(path,"test-sample",wb, ws5)[0], 133)
		self.assertEqual(get_alignment_metrics_rmdup(path,"test-sample",wb, ws5)[1], 12.337662337662337)


	def test_get_subpanel_NTC_check(self):

		aligned_reads_value_rmdup=133

		wb=Workbook()
		ws7=wb.create_sheet("Patient_demographics")
		ws6= wb.create_sheet("NTC fusion report")
		ws9=wb.create_sheet("Subpanel_NTC_Check")
		ws1= wb.create_sheet("Gene_fusion_report")
		ws5=wb.create_sheet("Summary")
		ws3=wb.create_sheet("coverage_with_duplicates")
		ws4=wb.create_sheet("coverage_without_duplicates")
		ws2=wb.create_sheet("total_coverage")

		ntc_average_depth=(ntc_get_coverage(path, "test-NTC"))[0]
		ntc_average_depth_rmdup=(ntc_get_coverage(path, "test-NTC"))[1]

		coverage_rmdup=get_coverage("Lung", ntc_average_depth, ntc_average_depth_rmdup, "test-sample", path,wb, ws3, ws4)[0]

		coverage=get_coverage("Lung", ntc_average_depth, ntc_average_depth_rmdup, "test-sample", path,wb, ws3, ws4)[1]

		#Colorectal
		wb_Colorectal3=Workbook()
		ws7_Colorectal3=wb_Colorectal3.create_sheet("Patient_demographics")
		ws6_Colorectal3= wb_Colorectal3.create_sheet("NTC fusion report")
		ws9_Colorectal3=wb_Colorectal3.create_sheet("Subpanel_NTC_Check")
		ws1_Colorectal3= wb_Colorectal3.create_sheet("Gene_fusion_report")
		ws5_Colorectal3=wb_Colorectal3.create_sheet("Summary")
		ws3_Colorectal3=wb_Colorectal3.create_sheet("coverage_with_duplicates")
		ws4_Colorectal3=wb_Colorectal3.create_sheet("coverage_without_duplicates")
		ws2_Colorectal3=wb_Colorectal3.create_sheet("total_coverage")	

		ws9=(get_subpanel_NTC_check(coverage, aligned_reads_value_rmdup, path, "test-sample","test-NTC", referral_list_Colorectal, coverage_rmdup,wb_Colorectal3, ws9_Colorectal3)[0])


		self.assertEqual(ws9["B15"].value,133)
		self.assertEqual(ws9["C15"].value,4)
		self.assertEqual(ws9["D15"].value,3.007518796992481)


		self.assertEqual(ws9["A24"].value,None)
		self.assertEqual(ws9["B24"].value,None)
		self.assertEqual(ws9["C24"].value,None)
		self.assertEqual(ws9["D24"].value,None)


		#GIST
		wb_GIST3=Workbook()
		ws7_GIST3=wb_GIST3.create_sheet("Patient_demographics")
		ws6_GIST3= wb_GIST3.create_sheet("NTC fusion report")
		ws9_GIST3=wb_GIST3.create_sheet("Subpanel_NTC_Check")
		ws1_GIST3= wb_GIST3.create_sheet("Gene_fusion_report")
		ws5_GIST3=wb_GIST3.create_sheet("Summary")
		ws3_GIST3=wb_GIST3.create_sheet("coverage_with_duplicates")
		ws4_GIST3=wb_GIST3.create_sheet("coverage_without_duplicates")
		ws2_GIST3=wb_GIST3.create_sheet("total_coverage")	

		ws9=(get_subpanel_NTC_check(coverage, aligned_reads_value_rmdup, path, "test-sample","test-NTC", referral_list_GIST, coverage_rmdup,wb_GIST3, ws9_GIST3)[0])

		self.assertEqual(ws9["B15"].value,133)
		self.assertEqual(ws9["C15"].value,4)
		self.assertEqual(ws9["D15"].value,3.007518796992481)


		self.assertEqual(ws9["A24"].value,None)
		self.assertEqual(ws9["B24"].value,None)
		self.assertEqual(ws9["C24"].value,None)
		self.assertEqual(ws9["D24"].value,None)




		#Glioma
		wb_Glioma3=Workbook()
		ws7_Glioma3=wb_Glioma3.create_sheet("Patient_demographics")
		ws6_Glioma3= wb_Glioma3.create_sheet("NTC fusion report")
		ws9_Glioma3=wb_Glioma3.create_sheet("Subpanel_NTC_Check")
		ws1_Glioma3= wb_Glioma3.create_sheet("Gene_fusion_report")
		ws5_Glioma3=wb_Glioma3.create_sheet("Summary")
		ws3_Glioma3=wb_Glioma3.create_sheet("coverage_with_duplicates")
		ws4_Glioma3=wb_Glioma3.create_sheet("coverage_without_duplicates")
		ws2_Glioma3=wb_Glioma3.create_sheet("total_coverage")	

		ws9=(get_subpanel_NTC_check(coverage, aligned_reads_value_rmdup, path, "test-sample","test-NTC", referral_list_Glioma, coverage_rmdup,wb_Glioma3, ws9_Glioma3)[0])

		self.assertEqual(ws9["B15"].value,133)
		self.assertEqual(ws9["C15"].value,4)
		self.assertEqual(ws9["D15"].value,3.007518796992481)


		self.assertEqual(ws9["A24"].value,"EGFR_Exon1_last10bases")
		self.assertEqual(ws9["B24"].value,2.0)
		self.assertEqual(ws9["C24"].value,1.0)
		self.assertEqual(ws9["D24"].value,50.0)


		self.assertEqual(ws9["A25"].value,"EGFR_Exon8_first10bases")
		self.assertEqual(ws9["B25"].value,9.0)
		self.assertEqual(ws9["C25"].value,1.0)
		self.assertEqual(ws9["D25"].value,11.11111111111111)

		self.assertEqual(ws9["A26"].value,None)
		self.assertEqual(ws9["B26"].value,None)
		self.assertEqual(ws9["C26"].value,None)
		self.assertEqual(ws9["D26"].value,None)




		#Lung
		wb_Lung3=Workbook()
		ws7_Lung3=wb_Lung3.create_sheet("Patient_demographics")
		ws6_Lung3= wb_Lung3.create_sheet("NTC fusion report")
		ws9_Lung3=wb_Lung3.create_sheet("Subpanel_NTC_Check")
		ws1_Lung3= wb_Lung3.create_sheet("Gene_fusion_report")
		ws5_Lung3=wb_Lung3.create_sheet("Summary")
		ws3_Lung3=wb_Lung3.create_sheet("coverage_with_duplicates")
		ws4_Lung3=wb_Lung3.create_sheet("coverage_without_duplicates")
		ws2_Lung3=wb_Lung3.create_sheet("total_coverage")	

		ws9=(get_subpanel_NTC_check(coverage, aligned_reads_value_rmdup, path, "test-sample","test-NTC", referral_list_Lung, coverage_rmdup,wb_Lung3, ws9_Lung3)[0])

		self.assertEqual(ws9["B15"].value,133)
		self.assertEqual(ws9["C15"].value,4)
		self.assertEqual(ws9["D15"].value,3.007518796992481)


		self.assertEqual(ws9["A24"].value,"MET_Exon13_last10bases")
		self.assertEqual(ws9["B24"].value,3.0)
		self.assertEqual(ws9["C24"].value,2.0)
		self.assertEqual(ws9["D24"].value,66.66666666666666)


		self.assertEqual(ws9["A25"].value,"MET_Exon15_first10bases")
		self.assertEqual(ws9["B25"].value,5.0)
		self.assertEqual(ws9["C25"].value,2.0)
		self.assertEqual(ws9["D25"].value,40.0)

		self.assertEqual(ws9["A26"].value,"EGFR_Exon1_last10bases")
		self.assertEqual(ws9["B26"].value,2.0)
		self.assertEqual(ws9["C26"].value,1.0)
		self.assertEqual(ws9["D26"].value,50.0)

		self.assertEqual(ws9["A27"].value,"EGFR_Exon8_first10bases")
		self.assertEqual(ws9["B27"].value,9.0)
		self.assertEqual(ws9["C27"].value,1.0)
		self.assertEqual(ws9["D27"].value,11.11111111111111)

		self.assertEqual(ws9["A28"].value,None)
		self.assertEqual(ws9["B28"].value,None)
		self.assertEqual(ws9["C28"].value,None)
		self.assertEqual(ws9["D28"].value,None)



		#Melanoma
		wb_Melanoma3=Workbook()
		ws7_Melanoma3=wb_Melanoma3.create_sheet("Patient_demographics")
		ws6_Melanoma3= wb_Melanoma3.create_sheet("NTC fusion report")
		ws9_Melanoma3=wb_Melanoma3.create_sheet("Subpanel_NTC_Check")
		ws1_Melanoma3= wb_Melanoma3.create_sheet("Gene_fusion_report")
		ws5_Melanoma3=wb_Melanoma3.create_sheet("Summary")
		ws3_Melanoma3=wb_Melanoma3.create_sheet("coverage_with_duplicates")
		ws4_Melanoma3=wb_Melanoma3.create_sheet("coverage_without_duplicates")
		ws2_Melanoma3=wb_Melanoma3.create_sheet("total_coverage")	

		ws9=(get_subpanel_NTC_check(coverage, aligned_reads_value_rmdup, path, "test-sample","test-NTC", referral_list_Melanoma, coverage_rmdup,wb_Melanoma3, ws9_Melanoma3)[0])

		self.assertEqual(ws9["B15"].value,133)
		self.assertEqual(ws9["C15"].value,4)
		self.assertEqual(ws9["D15"].value,3.007518796992481)


		self.assertEqual(ws9["A24"].value,None)
		self.assertEqual(ws9["B24"].value,None)
		self.assertEqual(ws9["C24"].value,None)
		self.assertEqual(ws9["D24"].value,None)


		self.assertEqual(ws9["A25"].value,None)
		self.assertEqual(ws9["B25"].value,None)
		self.assertEqual(ws9["C25"].value,None)
		self.assertEqual(ws9["D25"].value,None)

		#NTRK
		wb_NTRK3=Workbook()
		ws7_NTRK3=wb_NTRK3.create_sheet("Patient_demographics")
		ws6_NTRK3= wb_NTRK3.create_sheet("NTC fusion report")
		ws9_NTRK3=wb_NTRK3.create_sheet("Subpanel_NTC_Check")
		ws1_NTRK3= wb_NTRK3.create_sheet("Gene_fusion_report")
		ws5_NTRK3=wb_NTRK3.create_sheet("Summary")
		ws3_NTRK3=wb_NTRK3.create_sheet("coverage_with_duplicates")
		ws4_NTRK3=wb_NTRK3.create_sheet("coverage_without_duplicates")
		ws2_NTRK3=wb_NTRK3.create_sheet("total_coverage")	

		ws9=(get_subpanel_NTC_check(coverage, aligned_reads_value_rmdup, path, "test-sample","test-NTC", referral_list_NTRK, coverage_rmdup,wb_NTRK3, ws9_NTRK3)[0])
		self.assertEqual(ws9["B15"].value,133)
		self.assertEqual(ws9["C15"].value,4)
		self.assertEqual(ws9["D15"].value,3.007518796992481)


		self.assertEqual(ws9["A24"].value,None)
		self.assertEqual(ws9["B24"].value,None)
		self.assertEqual(ws9["C24"].value,None)
		self.assertEqual(ws9["D24"].value,None)


		self.assertEqual(ws9["A25"].value,None)
		self.assertEqual(ws9["B25"].value,None)
		self.assertEqual(ws9["C25"].value,None)
		self.assertEqual(ws9["D25"].value,None)

		#Thyroid

		wb_Thyroid3=Workbook()
		ws7_Thyroid3=wb_Thyroid3.create_sheet("Patient_demographics")
		ws6_Thyroid3= wb_Thyroid3.create_sheet("NTC fusion report")
		ws9_Thyroid3=wb_Thyroid3.create_sheet("Subpanel_NTC_Check")
		ws1_Thyroid3= wb_Thyroid3.create_sheet("Gene_fusion_report")
		ws5_Thyroid3=wb_Thyroid3.create_sheet("Summary")
		ws3_Thyroid3=wb_Thyroid3.create_sheet("coverage_with_duplicates")
		ws4_Thyroid3=wb_Thyroid3.create_sheet("coverage_without_duplicates")
		ws2_Thyroid3=wb_Thyroid3.create_sheet("total_coverage")	

		ws9=(get_subpanel_NTC_check(coverage, aligned_reads_value_rmdup, path, "test-sample","test-NTC", referral_list_Thyroid, coverage_rmdup,wb_Thyroid3, ws9_Thyroid3)[0])

		self.assertEqual(ws9["B15"].value,133)
		self.assertEqual(ws9["C15"].value,4)
		self.assertEqual(ws9["D15"].value,3.007518796992481)


		self.assertEqual(ws9["A24"].value,"EGFR_Exon1_last10bases")
		self.assertEqual(ws9["B24"].value,2.0)
		self.assertEqual(ws9["C24"].value,1.0)
		self.assertEqual(ws9["D24"].value,50.0)


		self.assertEqual(ws9["A25"].value,"EGFR_Exon8_first10bases")
		self.assertEqual(ws9["B25"].value,9.0)
		self.assertEqual(ws9["C25"].value,1.0)
		self.assertEqual(ws9["D25"].value,11.11111111111111)

		self.assertEqual(ws9["A26"].value,None)
		self.assertEqual(ws9["B26"].value,None)
		self.assertEqual(ws9["C26"].value,None)
		self.assertEqual(ws9["D26"].value,None)

		#Tumour
		wb_Tumour3=Workbook()
		ws7_Tumour3=wb_Tumour3.create_sheet("Patient_demographics")
		ws6_Tumour3= wb_Tumour3.create_sheet("NTC fusion report")
		ws9_Tumour3=wb_Tumour3.create_sheet("Subpanel_NTC_Check")
		ws1_Tumour3= wb_Tumour3.create_sheet("Gene_fusion_report")
		ws5_Tumour3=wb_Tumour3.create_sheet("Summary")
		ws3_Tumour3=wb_Tumour3.create_sheet("coverage_with_duplicates")
		ws4_Tumour3=wb_Tumour3.create_sheet("coverage_without_duplicates")
		ws2_Tumour3=wb_Tumour3.create_sheet("total_coverage")	

		ws9=(get_subpanel_NTC_check(coverage, aligned_reads_value_rmdup, path, "test-sample","test-NTC", referral_list_Tumour, coverage_rmdup,wb_Tumour3, ws9_Tumour3)[0])

		self.assertEqual(ws9["B15"].value,133)
		self.assertEqual(ws9["C15"].value,4)
		self.assertEqual(ws9["D15"].value,3.007518796992481)

		self.assertEqual(ws9["A24"].value,"MET_Exon13_last10bases")
		self.assertEqual(ws9["B24"].value,3.0)
		self.assertEqual(ws9["C24"].value,2.0)
		self.assertEqual(ws9["D24"].value,66.66666666666666)

		self.assertEqual(ws9["A25"].value,"MET_Exon15_first10bases")
		self.assertEqual(ws9["B25"].value,5.0)
		self.assertEqual(ws9["C25"].value,2.0)
		self.assertEqual(ws9["D25"].value,40.0)

		self.assertEqual(ws9["A26"].value,"EGFR_Exon1_last10bases")
		self.assertEqual(ws9["B26"].value,2.0)
		self.assertEqual(ws9["C26"].value,1.0)
		self.assertEqual(ws9["D26"].value,50.0)


		self.assertEqual(ws9["A27"].value,"EGFR_Exon8_first10bases")
		self.assertEqual(ws9["B27"].value,9.0)
		self.assertEqual(ws9["C27"].value,1.0)
		self.assertEqual(ws9["D27"].value,11.11111111111111)

		self.assertEqual(ws9["A28"].value,None)
		self.assertEqual(ws9["B28"].value,None)
		self.assertEqual(ws9["C28"].value,None)
		self.assertEqual(ws9["D28"].value,None)




	def test_get_met_exon_skipping(self):

		wb=Workbook()
		ws7=wb.create_sheet("Patient_demographics")
		ws6= wb.create_sheet("NTC fusion report")
		ws9=wb.create_sheet("Subpanel_NTC_Check")
		ws1= wb.create_sheet("Gene_fusion_report")
		ws5=wb.create_sheet("Summary")
		ws3=wb.create_sheet("coverage_with_duplicates")
		ws4=wb.create_sheet("coverage_without_duplicates")
		ws2=wb.create_sheet("total_coverage")

		ntc_total_coverage=pandas.read_csv(path+"test-NTC"+"/"+"test-NTC"+"_coverage.totalCoverage", sep="\t")
		ntc_total_average_depth=ntc_total_coverage["AVG_DEPTH"]

		ntc_total_coverage_rmdup=pandas.read_csv(path+"test-NTC"+"/"+"test-NTC"+"_rmdup_coverage.totalCoverage", sep="\t")
		ntc_total_average_depth_rmdup=ntc_total_coverage_rmdup["AVG_DEPTH"]

		ntc_average_depth=(ntc_get_coverage(path, "test-NTC"))[0]
		ntc_average_depth_rmdup=(ntc_get_coverage(path, "test-NTC"))[1]

		coverage_rmdup=get_coverage("Colorectal", ntc_average_depth, ntc_average_depth_rmdup, "test-sample", path, wb, ws3, ws4)[0]

		#Colorectal
		wb_Colorectal_test=Workbook()
		ws7_Colorectal_test=wb_Colorectal_test.create_sheet("Patient_demographics")
		ws6_Colorectal_test= wb_Colorectal_test.create_sheet("NTC fusion report")
		ws9_Colorectal_test=wb_Colorectal_test.create_sheet("Subpanel_NTC_Check")
		ws1_Colorectal_test= wb_Colorectal_test.create_sheet("Gene_fusion_report")
		ws5_Colorectal_test=wb_Colorectal_test.create_sheet("Summary")
		ws3_Colorectal_test=wb_Colorectal_test.create_sheet("coverage_with_duplicates")
		ws4_Colorectal_test=wb_Colorectal_test.create_sheet("coverage_without_duplicates")
		ws2_Colorectal_test=wb_Colorectal_test.create_sheet("total_coverage")	

		wb_Colorectal_test2=Workbook()
		ws7_Colorectal_test2=wb_Colorectal_test2.create_sheet("Patient_demographics")
		ws6_Colorectal_test2= wb_Colorectal_test2.create_sheet("NTC fusion report")
		ws9_Colorectal_test2=wb_Colorectal_test2.create_sheet("Subpanel_NTC_Check")
		ws1_Colorectal_test2= wb_Colorectal_test2.create_sheet("Gene_fusion_report")
		ws5_Colorectal_test2=wb_Colorectal_test2.create_sheet("Summary")
		ws3_Colorectal_test2=wb_Colorectal_test2.create_sheet("coverage_with_duplicates")
		ws4_Colorectal_test2=wb_Colorectal_test2.create_sheet("coverage_without_duplicates")
		ws2_Colorectal_test2=wb_Colorectal_test2.create_sheet("total_coverage")

		wb_Colorectal_test3=Workbook()
		ws7_Colorectal_test3=wb_Colorectal_test3.create_sheet("Patient_demographics")
		ws6_Colorectal_test3= wb_Colorectal_test3.create_sheet("NTC fusion report")
		ws9_Colorectal_test3=wb_Colorectal_test3.create_sheet("Subpanel_NTC_Check")
		ws1_Colorectal_test3= wb_Colorectal_test3.create_sheet("Gene_fusion_report")
		ws5_Colorectal_test3=wb_Colorectal_test3.create_sheet("Summary")
		ws3_Colorectal_test3=wb_Colorectal_test3.create_sheet("coverage_with_duplicates")
		ws4_Colorectal_test3=wb_Colorectal_test3.create_sheet("coverage_without_duplicates")
		ws2_Colorectal_test3=wb_Colorectal_test3.create_sheet("total_coverage")

		#test rmats results for Colorectal referral type
		#test_sample1
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Colorectal, "test-sample", "Colorectal", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Colorectal_test, ws5_Colorectal_test, ws6_Colorectal_test))
		self.assertEqual(ws8["B10"].value,None)
		self.assertEqual(ws8["B11"].value,None)

		#test_sample2
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Colorectal, "test-sample2", "Colorectal", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Colorectal_test2, ws5_Colorectal_test2, ws6_Colorectal_test2))
		self.assertEqual(ws8["B10"].value,None)
		self.assertEqual(ws8["B11"].value,None)

		#test_sample3
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Colorectal, "test-sample3", "Colorectal", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Colorectal_test3, ws5_Colorectal_test3, ws6_Colorectal_test3))
		self.assertEqual(ws8["B10"].value, None)
		self.assertEqual(ws8["B11"].value,None)

		#GIST
		wb_GIST_test=Workbook()
		ws7_GIST_test=wb_GIST_test.create_sheet("Patient_demographics")
		ws6_GIST_test= wb_GIST_test.create_sheet("NTC fusion report")
		ws9_GIST_test=wb_GIST_test.create_sheet("Subpanel_NTC_Check")
		ws1_GIST_test= wb_GIST_test.create_sheet("Gene_fusion_report")
		ws5_GIST_test=wb_GIST_test.create_sheet("Summary")
		ws3_GIST_test=wb_GIST_test.create_sheet("coverage_with_duplicates")
		ws4_GIST_test=wb_GIST_test.create_sheet("coverage_without_duplicates")
		ws2_GIST_test=wb_GIST_test.create_sheet("total_coverage")	

		wb_GIST_test2=Workbook()
		ws7_GIST_test2=wb_GIST_test2.create_sheet("Patient_demographics")
		ws6_GIST_test2= wb_GIST_test2.create_sheet("NTC fusion report")
		ws9_GIST_test2=wb_GIST_test2.create_sheet("Subpanel_NTC_Check")
		ws1_GIST_test2= wb_GIST_test2.create_sheet("Gene_fusion_report")
		ws5_GIST_test2=wb_GIST_test2.create_sheet("Summary")
		ws3_GIST_test2=wb_GIST_test2.create_sheet("coverage_with_duplicates")
		ws4_GIST_test2=wb_GIST_test2.create_sheet("coverage_without_duplicates")
		ws2_GIST_test2=wb_GIST_test2.create_sheet("total_coverage")

		wb_GIST_test3=Workbook()
		ws7_GIST_test3=wb_GIST_test3.create_sheet("Patient_demographics")
		ws6_GIST_test3= wb_GIST_test3.create_sheet("NTC fusion report")
		ws9_GIST_test3=wb_GIST_test3.create_sheet("Subpanel_NTC_Check")
		ws1_GIST_test3= wb_GIST_test3.create_sheet("Gene_fusion_report")
		ws5_GIST_test3=wb_GIST_test3.create_sheet("Summary")
		ws3_GIST_test3=wb_GIST_test3.create_sheet("coverage_with_duplicates")
		ws4_GIST_test3=wb_GIST_test3.create_sheet("coverage_without_duplicates")
		ws2_GIST_test3=wb_GIST_test3.create_sheet("total_coverage")

		#test rmats results for GIST referral type
		#test_sample1
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_GIST, "test-sample", "GIST", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_GIST_test, ws5_GIST_test, ws6_GIST_test))
		self.assertEqual(ws8["B10"].value,None)
		self.assertEqual(ws8["B11"].value,None)

		#test_sample2
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_GIST, "test-sample2", "GIST", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_GIST_test2, ws5_GIST_test2, ws6_GIST_test2))
		self.assertEqual(ws8["B10"].value,None)
		self.assertEqual(ws8["B11"].value,None)

		#test_sample3
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_GIST, "test-sample3", "GIST", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_GIST_test3, ws5_GIST_test3, ws6_GIST_test3))
		self.assertEqual(ws8["B10"].value, None)
		self.assertEqual(ws8["B11"].value,None)

		#Glioma
		wb_Glioma_test=Workbook()
		ws7_Glioma_test=wb_Glioma_test.create_sheet("Patient_demographics")
		ws6_Glioma_test= wb_Glioma_test.create_sheet("NTC fusion report")
		ws9_Glioma_test=wb_Glioma_test.create_sheet("Subpanel_NTC_Check")
		ws1_Glioma_test= wb_Glioma_test.create_sheet("Gene_fusion_report")
		ws5_Glioma_test=wb_Glioma_test.create_sheet("Summary")
		ws3_Glioma_test=wb_Glioma_test.create_sheet("coverage_with_duplicates")
		ws4_Glioma_test=wb_Glioma_test.create_sheet("coverage_without_duplicates")
		ws2_Glioma_test=wb_Glioma_test.create_sheet("total_coverage")	

		wb_Glioma_test2=Workbook()
		ws7_Glioma_test2=wb_Glioma_test2.create_sheet("Patient_demographics")
		ws6_Glioma_test2= wb_Glioma_test2.create_sheet("NTC fusion report")
		ws9_Glioma_test2=wb_Glioma_test2.create_sheet("Subpanel_NTC_Check")
		ws1_Glioma_test2= wb_Glioma_test2.create_sheet("Gene_fusion_report")
		ws5_Glioma_test2=wb_Glioma_test2.create_sheet("Summary")
		ws3_Glioma_test2=wb_Glioma_test2.create_sheet("coverage_with_duplicates")
		ws4_Glioma_test2=wb_Glioma_test2.create_sheet("coverage_without_duplicates")
		ws2_Glioma_test2=wb_Glioma_test2.create_sheet("total_coverage")

		wb_Glioma_test3=Workbook()
		ws7_Glioma_test3=wb_Glioma_test3.create_sheet("Patient_demographics")
		ws6_Glioma_test3= wb_Glioma_test3.create_sheet("NTC fusion report")
		ws9_Glioma_test3=wb_Glioma_test3.create_sheet("Subpanel_NTC_Check")
		ws1_Glioma_test3= wb_Glioma_test3.create_sheet("Gene_fusion_report")
		ws5_Glioma_test3=wb_Glioma_test3.create_sheet("Summary")
		ws3_Glioma_test3=wb_Glioma_test3.create_sheet("coverage_with_duplicates")
		ws4_Glioma_test3=wb_Glioma_test3.create_sheet("coverage_without_duplicates")
		ws2_Glioma_test3=wb_Glioma_test3.create_sheet("total_coverage")

		#test rmats results for Glioma referral type
		#test_sample1
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Glioma, "test-sample", "Glioma", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Glioma_test, ws5_Glioma_test, ws6_Glioma_test))
		self.assertEqual(ws8["B10"].value,"EGFRv3-no fusions found")
		self.assertEqual(ws8["B11"].value,None)

		#test_sample2
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Glioma, "test-sample2", "Glioma", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Glioma_test2, ws5_Glioma_test2, ws6_Glioma_test2))
		self.assertEqual(ws8["B10"].value,'"EGFR"')
		self.assertEqual(ws8["B11"].value,None)

		#test_sample3
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Glioma, "test-sample3", "Glioma", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Glioma_test3, ws5_Glioma_test3, ws6_Glioma_test3))
		self.assertEqual(ws8["B10"].value,'"EGFR"')
		self.assertEqual(ws8["B11"].value,None)

		#Lung
		wb_Lung_test=Workbook()
		ws7_Lung_test=wb_Lung_test.create_sheet("Patient_demographics")
		ws6_Lung_test= wb_Lung_test.create_sheet("NTC fusion report")
		ws9_Lung_test=wb_Lung_test.create_sheet("Subpanel_NTC_Check")
		ws1_Lung_test= wb_Lung_test.create_sheet("Gene_fusion_report")
		ws5_Lung_test=wb_Lung_test.create_sheet("Summary")
		ws3_Lung_test=wb_Lung_test.create_sheet("coverage_with_duplicates")
		ws4_Lung_test=wb_Lung_test.create_sheet("coverage_without_duplicates")
		ws2_Lung_test=wb_Lung_test.create_sheet("total_coverage")	

		wb_Lung_test2=Workbook()
		ws7_Lung_test2=wb_Lung_test2.create_sheet("Patient_demographics")
		ws6_Lung_test2= wb_Lung_test2.create_sheet("NTC fusion report")
		ws9_Lung_test2=wb_Lung_test2.create_sheet("Subpanel_NTC_Check")
		ws1_Lung_test2= wb_Lung_test2.create_sheet("Gene_fusion_report")
		ws5_Lung_test2=wb_Lung_test2.create_sheet("Summary")
		ws3_Lung_test2=wb_Lung_test2.create_sheet("coverage_with_duplicates")
		ws4_Lung_test2=wb_Lung_test2.create_sheet("coverage_without_duplicates")
		ws2_Lung_test2=wb_Lung_test2.create_sheet("total_coverage")

		wb_Lung_test3=Workbook()
		ws7_Lung_test3=wb_Lung_test3.create_sheet("Patient_demographics")
		ws6_Lung_test3= wb_Lung_test3.create_sheet("NTC fusion report")
		ws9_Lung_test3=wb_Lung_test3.create_sheet("Subpanel_NTC_Check")
		ws1_Lung_test3= wb_Lung_test3.create_sheet("Gene_fusion_report")
		ws5_Lung_test3=wb_Lung_test3.create_sheet("Summary")
		ws3_Lung_test3=wb_Lung_test3.create_sheet("coverage_with_duplicates")
		ws4_Lung_test3=wb_Lung_test3.create_sheet("coverage_without_duplicates")
		ws2_Lung_test3=wb_Lung_test3.create_sheet("total_coverage")

		#test rmats results for Lung referral type
		#test_sample1
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Lung, "test-sample", "Lung", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Lung_test, ws5_Lung_test, ws6_Lung_test))
		self.assertEqual(ws8["B10"].value,"EGFRv3-no fusions found")
		self.assertEqual(ws8["B11"].value,'"MET"' )

		#test_sample2
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Lung, "test-sample2", "Lung", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Lung_test2, ws5_Lung_test2, ws6_Lung_test2))
		self.assertEqual(ws8["B10"].value,'"EGFR"')
		self.assertEqual(ws8["B11"].value,"MET_exon_14_skipping-no fusions found")

		#test_sample3
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Lung, "test-sample3", "Lung", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Lung_test3, ws5_Lung_test3, ws6_Lung_test3))
		self.assertEqual(ws8["B10"].value, '"EGFR"')
		self.assertEqual(ws8["B11"].value,'"MET"')

		#Melanoma
		wb_Melanoma_test=Workbook()
		ws7_Melanoma_test=wb_Melanoma_test.create_sheet("Patient_demographics")
		ws6_Melanoma_test= wb_Melanoma_test.create_sheet("NTC fusion report")
		ws9_Melanoma_test=wb_Melanoma_test.create_sheet("Subpanel_NTC_Check")
		ws1_Melanoma_test= wb_Melanoma_test.create_sheet("Gene_fusion_report")
		ws5_Melanoma_test=wb_Melanoma_test.create_sheet("Summary")
		ws3_Melanoma_test=wb_Melanoma_test.create_sheet("coverage_with_duplicates")
		ws4_Melanoma_test=wb_Melanoma_test.create_sheet("coverage_without_duplicates")
		ws2_Melanoma_test=wb_Melanoma_test.create_sheet("total_coverage")	

		wb_Melanoma_test2=Workbook()
		ws7_Melanoma_test2=wb_Melanoma_test2.create_sheet("Patient_demographics")
		ws6_Melanoma_test2= wb_Melanoma_test2.create_sheet("NTC fusion report")
		ws9_Melanoma_test2=wb_Melanoma_test2.create_sheet("Subpanel_NTC_Check")
		ws1_Melanoma_test2= wb_Melanoma_test2.create_sheet("Gene_fusion_report")
		ws5_Melanoma_test2=wb_Melanoma_test2.create_sheet("Summary")
		ws3_Melanoma_test2=wb_Melanoma_test2.create_sheet("coverage_with_duplicates")
		ws4_Melanoma_test2=wb_Melanoma_test2.create_sheet("coverage_without_duplicates")
		ws2_Melanoma_test2=wb_Melanoma_test2.create_sheet("total_coverage")

		wb_Melanoma_test3=Workbook()
		ws7_Melanoma_test3=wb_Melanoma_test3.create_sheet("Patient_demographics")
		ws6_Melanoma_test3= wb_Melanoma_test3.create_sheet("NTC fusion report")
		ws9_Melanoma_test3=wb_Melanoma_test3.create_sheet("Subpanel_NTC_Check")
		ws1_Melanoma_test3= wb_Melanoma_test3.create_sheet("Gene_fusion_report")
		ws5_Melanoma_test3=wb_Melanoma_test3.create_sheet("Summary")
		ws3_Melanoma_test3=wb_Melanoma_test3.create_sheet("coverage_with_duplicates")
		ws4_Melanoma_test3=wb_Melanoma_test3.create_sheet("coverage_without_duplicates")
		ws2_Melanoma_test3=wb_Melanoma_test3.create_sheet("total_coverage")

		#test rmats results for Melanoma referral type
		#test_sample1
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Melanoma, "test-sample", "Melanoma", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Melanoma_test, ws5_Melanoma_test, ws6_Melanoma_test))
		self.assertEqual(ws8["B10"].value,None)
		self.assertEqual(ws8["B11"].value,None)

		#test_sample2
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Melanoma, "test-sample2", "Melanoma", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Melanoma_test2, ws5_Melanoma_test2, ws6_Melanoma_test2))
		self.assertEqual(ws8["B10"].value,None)
		self.assertEqual(ws8["B11"].value,None)

		#test_sample3
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Melanoma, "test-sample3", "Melanoma", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Melanoma_test3, ws5_Melanoma_test3, ws6_Melanoma_test3))
		self.assertEqual(ws8["B10"].value, None)
		self.assertEqual(ws8["B11"].value,None)

		#NTRK
		wb_NTRK_test=Workbook()
		ws7_NTRK_test=wb_NTRK_test.create_sheet("Patient_demographics")
		ws6_NTRK_test= wb_NTRK_test.create_sheet("NTC fusion report")
		ws9_NTRK_test=wb_NTRK_test.create_sheet("Subpanel_NTC_Check")
		ws1_NTRK_test= wb_NTRK_test.create_sheet("Gene_fusion_report")
		ws5_NTRK_test=wb_NTRK_test.create_sheet("Summary")
		ws3_NTRK_test=wb_NTRK_test.create_sheet("coverage_with_duplicates")
		ws4_NTRK_test=wb_NTRK_test.create_sheet("coverage_without_duplicates")
		ws2_NTRK_test=wb_NTRK_test.create_sheet("total_coverage")	

		wb_NTRK_test2=Workbook()
		ws7_NTRK_test2=wb_NTRK_test2.create_sheet("Patient_demographics")
		ws6_NTRK_test2= wb_NTRK_test2.create_sheet("NTC fusion report")
		ws9_NTRK_test2=wb_NTRK_test2.create_sheet("Subpanel_NTC_Check")
		ws1_NTRK_test2= wb_NTRK_test2.create_sheet("Gene_fusion_report")
		ws5_NTRK_test2=wb_NTRK_test2.create_sheet("Summary")
		ws3_NTRK_test2=wb_NTRK_test2.create_sheet("coverage_with_duplicates")
		ws4_NTRK_test2=wb_NTRK_test2.create_sheet("coverage_without_duplicates")
		ws2_NTRK_test2=wb_NTRK_test2.create_sheet("total_coverage")

		wb_NTRK_test3=Workbook()
		ws7_NTRK_test3=wb_NTRK_test3.create_sheet("Patient_demographics")
		ws6_NTRK_test3= wb_NTRK_test3.create_sheet("NTC fusion report")
		ws9_NTRK_test3=wb_NTRK_test3.create_sheet("Subpanel_NTC_Check")
		ws1_NTRK_test3= wb_NTRK_test3.create_sheet("Gene_fusion_report")
		ws5_NTRK_test3=wb_NTRK_test3.create_sheet("Summary")
		ws3_NTRK_test3=wb_NTRK_test3.create_sheet("coverage_with_duplicates")
		ws4_NTRK_test3=wb_NTRK_test3.create_sheet("coverage_without_duplicates")
		ws2_NTRK_test3=wb_NTRK_test3.create_sheet("total_coverage")

		#test rmats results for Glioma referral type
		#test_sample1
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_NTRK, "test-sample", "NTRK", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_NTRK_test, ws5_NTRK_test, ws6_NTRK_test))
		self.assertEqual(ws8["B10"].value,None)
		self.assertEqual(ws8["B11"].value,None)

		#test_sample2
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_NTRK, "test-sample2", "NTRK", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_NTRK_test2, ws5_NTRK_test2, ws6_NTRK_test2))
		self.assertEqual(ws8["B10"].value,None)
		self.assertEqual(ws8["B11"].value,None)

		#test_sample3
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_NTRK, "test-sample3", "NTRK", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_NTRK_test3, ws5_NTRK_test3, ws6_NTRK_test3))
		self.assertEqual(ws8["B10"].value, None)
		self.assertEqual(ws8["B11"].value,None)

		#Thyroid
		wb_Thyroid_test=Workbook()
		ws7_Thyroid_test=wb_Thyroid_test.create_sheet("Patient_demographics")
		ws6_Thyroid_test= wb_Thyroid_test.create_sheet("NTC fusion report")
		ws9_Thyroid_test=wb_Thyroid_test.create_sheet("Subpanel_NTC_Check")
		ws1_Thyroid_test= wb_Thyroid_test.create_sheet("Gene_fusion_report")
		ws5_Thyroid_test=wb_Thyroid_test.create_sheet("Summary")
		ws3_Thyroid_test=wb_Thyroid_test.create_sheet("coverage_with_duplicates")
		ws4_Thyroid_test=wb_Thyroid_test.create_sheet("coverage_without_duplicates")
		ws2_Thyroid_test=wb_Thyroid_test.create_sheet("total_coverage")	

		wb_Thyroid_test2=Workbook()
		ws7_Thyroid_test2=wb_Thyroid_test2.create_sheet("Patient_demographics")
		ws6_Thyroid_test2= wb_Thyroid_test2.create_sheet("NTC fusion report")
		ws9_Thyroid_test2=wb_Thyroid_test2.create_sheet("Subpanel_NTC_Check")
		ws1_Thyroid_test2= wb_Thyroid_test2.create_sheet("Gene_fusion_report")
		ws5_Thyroid_test2=wb_Thyroid_test2.create_sheet("Summary")
		ws3_Thyroid_test2=wb_Thyroid_test2.create_sheet("coverage_with_duplicates")
		ws4_Thyroid_test2=wb_Thyroid_test2.create_sheet("coverage_without_duplicates")
		ws2_Thyroid_test2=wb_Thyroid_test2.create_sheet("total_coverage")

		wb_Thyroid_test3=Workbook()
		ws7_Thyroid_test3=wb_Thyroid_test3.create_sheet("Patient_demographics")
		ws6_Thyroid_test3= wb_Thyroid_test3.create_sheet("NTC fusion report")
		ws9_Thyroid_test3=wb_Thyroid_test3.create_sheet("Subpanel_NTC_Check")
		ws1_Thyroid_test3= wb_Thyroid_test3.create_sheet("Gene_fusion_report")
		ws5_Thyroid_test3=wb_Thyroid_test3.create_sheet("Summary")
		ws3_Thyroid_test3=wb_Thyroid_test3.create_sheet("coverage_with_duplicates")
		ws4_Thyroid_test3=wb_Thyroid_test3.create_sheet("coverage_without_duplicates")
		ws2_Thyroid_test3=wb_Thyroid_test3.create_sheet("total_coverage")

		#test rmats results for Thyroid referral type
		#test_sample1
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Thyroid, "test-sample", "Thyroid", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Thyroid_test, ws5_Thyroid_test, ws6_Thyroid_test))
		self.assertEqual(ws8["B10"].value,"EGFRv3-no fusions found")
		self.assertEqual(ws8["B11"].value,None)

		#test_sample2
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Thyroid, "test-sample2", "Thyroid", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Thyroid_test2, ws5_Thyroid_test2, ws6_Thyroid_test2))
		self.assertEqual(ws8["B10"].value,'"EGFR"')
		self.assertEqual(ws8["B11"].value,None)

		#test_sample3
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Thyroid, "test-sample3", "Thyroid", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Thyroid_test3, ws5_Thyroid_test3, ws6_Thyroid_test3))
		self.assertEqual(ws8["B10"].value,'"EGFR"')
		self.assertEqual(ws8["B11"].value,None)

		#Tumour
		wb_Tumour_test=Workbook()
		ws7_Tumour_test=wb_Tumour_test.create_sheet("Patient_demographics")
		ws6_Tumour_test= wb_Tumour_test.create_sheet("NTC fusion report")
		ws9_Tumour_test=wb_Tumour_test.create_sheet("Subpanel_NTC_Check")
		ws1_Tumour_test= wb_Tumour_test.create_sheet("Gene_fusion_report")
		ws5_Tumour_test=wb_Tumour_test.create_sheet("Summary")
		ws3_Tumour_test=wb_Tumour_test.create_sheet("coverage_with_duplicates")
		ws4_Tumour_test=wb_Tumour_test.create_sheet("coverage_without_duplicates")
		ws2_Tumour_test=wb_Tumour_test.create_sheet("total_coverage")	

		wb_Tumour_test2=Workbook()
		ws7_Tumour_test2=wb_Tumour_test2.create_sheet("Patient_demographics")
		ws6_Tumour_test2= wb_Tumour_test2.create_sheet("NTC fusion report")
		ws9_Tumour_test2=wb_Tumour_test2.create_sheet("Subpanel_NTC_Check")
		ws1_Tumour_test2= wb_Tumour_test2.create_sheet("Gene_fusion_report")
		ws5_Tumour_test2=wb_Tumour_test2.create_sheet("Summary")
		ws3_Tumour_test2=wb_Tumour_test2.create_sheet("coverage_with_duplicates")
		ws4_Tumour_test2=wb_Tumour_test2.create_sheet("coverage_without_duplicates")
		ws2_Tumour_test2=wb_Tumour_test2.create_sheet("total_coverage")

		wb_Tumour_test3=Workbook()
		ws7_Tumour_test3=wb_Tumour_test3.create_sheet("Patient_demographics")
		ws6_Tumour_test3= wb_Tumour_test3.create_sheet("NTC fusion report")
		ws9_Tumour_test3=wb_Tumour_test3.create_sheet("Subpanel_NTC_Check")
		ws1_Tumour_test3= wb_Tumour_test3.create_sheet("Gene_fusion_report")
		ws5_Tumour_test3=wb_Tumour_test3.create_sheet("Summary")
		ws3_Tumour_test3=wb_Tumour_test3.create_sheet("coverage_with_duplicates")
		ws4_Tumour_test3=wb_Tumour_test3.create_sheet("coverage_without_duplicates")
		ws2_Tumour_test3=wb_Tumour_test3.create_sheet("total_coverage")

		#test rmats results for Tumour referral type
		#test_sample1
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Tumour, "test-sample", "Tumour", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Tumour_test, ws5_Tumour_test, ws6_Tumour_test))
		self.assertEqual(ws8["B10"].value,"EGFRv3-no fusions found")
		self.assertEqual(ws8["B11"].value,'"MET"')

		#test_sample2
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Tumour, "test-sample2", "Tumour", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Tumour_test2, ws5_Tumour_test2, ws6_Tumour_test2))
		self.assertEqual(ws8["B10"].value,'"EGFR"')
		self.assertEqual(ws8["B11"].value,"MET_exon_14_skipping-no fusions found")

		#test_sample3
		ws5, ws8, wb, ws6=(get_met_exon_skipping(referral_list_Tumour, "test-sample3", "Tumour", "worksheet", "validations", path, "test-NTC", coverage_rmdup,wb_Tumour_test3, ws5_Tumour_test3, ws6_Tumour_test3))
		self.assertEqual(ws8["B10"].value,'"EGFR"')
		self.assertEqual(ws8["B11"].value,'"MET"')


